"""
Validador detallado tabla 6063: CCAA + Sectores B-S + Jornada
Comparación completa INE CSV vs DuckDB
"""

import sys
from pathlib import Path
import pandas as pd
import duckdb
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración
sys.path.append(str(Path(__file__).parent))

def read_and_prepare_ine_csv(file_path):
    """
    Lee y prepara el CSV del INE tabla 6063
    """
    print("=" * 80)
    print("LEYENDO DATOS DEL INE - TABLA 6063")
    print("=" * 80)
    
    # Leer CSV con configuración correcta
    df = pd.read_csv(file_path, sep='\t', encoding='latin-1', decimal=',')
    
    print(f"\nArchivo: {file_path.name}")
    print(f"Columnas: {df.columns.tolist()}")
    print(f"Total registros: {len(df)}")
    
    # Filtrar solo 2025T1
    df_2025 = df[df['Periodo'] == '2025T1'].copy()
    print(f"Registros 2025T1: {len(df_2025)}")
    
    # Renombrar columna Total a Valor para consistencia
    df_2025.rename(columns={'Total': 'Valor'}, inplace=True)
    
    # Convertir valores a float
    df_2025['Valor'] = pd.to_numeric(df_2025['Valor'], errors='coerce')
    
    # Identificar columnas relevantes
    ccaa_col = None
    sector_col = None
    jornada_col = None
    
    for col in df_2025.columns:
        if 'Comunidades' in col or 'CCAA' in col or 'autónomas' in col.lower():
            ccaa_col = col
        elif 'Sectores' in col:
            sector_col = col
        elif 'jornada' in col.lower():
            jornada_col = col
    
    print(f"\nColumnas identificadas:")
    print(f"  - CCAA: {ccaa_col}")
    print(f"  - Sectores: {sector_col}")
    print(f"  - Jornada: {jornada_col}")
    
    # Mostrar resumen de datos
    if ccaa_col:
        print("\nAlgunas CCAA encontradas:")
        ccaas = df_2025[ccaa_col].unique()
        for i, ccaa in enumerate(ccaas[:5]):
            print(f"  - {ccaa}")
        if len(ccaas) > 5:
            print(f"  ... y {len(ccaas)-5} más")
        print(f"Total CCAA únicas: {len(ccaas)}")
    
    if jornada_col:
        print("\nTipos de jornada:")
        for jornada in df_2025[jornada_col].unique():
            count = len(df_2025[df_2025[jornada_col] == jornada])
            print(f"  - {jornada}: {count} registros")
    
    return df_2025, ccaa_col, sector_col, jornada_col

def get_duckdb_data_detailed(db_path):
    """
    Extrae datos detallados de DuckDB para tabla 6063
    """
    print("\n" + "=" * 80)
    print("EXTRAYENDO DATOS DE DUCKDB - TABLA 6063")
    print("=" * 80)
    
    conn = duckdb.connect(str(db_path))
    
    # Query completa para obtener todos los datos de 2025T1 tabla 6063
    query = """
    SELECT 
        tipo_jornada,
        cnae_nivel,
        cnae_codigo,
        cnae_nombre,
        ambito_territorial,
        ccaa_codigo,
        ccaa_nombre,
        metrica,
        causa,
        ROUND(valor, 1) as valor
    FROM observaciones_tiempo_trabajo
    WHERE fuente_tabla = '6063'
      AND periodo = '2025T1'
    ORDER BY ambito_territorial, ccaa_codigo, tipo_jornada, cnae_nivel, cnae_codigo, metrica
    """
    
    df = conn.execute(query).fetchdf()
    conn.close()
    
    print(f"\nTotal registros extraídos: {len(df)}")
    
    # Resumen por ámbito territorial
    print("\nResumen por ámbito territorial:")
    for ambito in df['ambito_territorial'].unique():
        count = len(df[df['ambito_territorial'] == ambito])
        print(f"  - {ambito}: {count} registros")
    
    # Resumen por CCAA
    print("\nCCAA en BD:")
    ccaas_db = df[df['ambito_territorial'] == 'CCAA']['ccaa_codigo'].unique()
    ccaas_db = sorted([c for c in ccaas_db if c is not None])
    print(f"  Total: {len(ccaas_db)} comunidades autónomas")
    for codigo in ccaas_db[:5]:
        nombre = df[df['ccaa_codigo'] == codigo]['ccaa_nombre'].iloc[0]
        print(f"  - {codigo}: {nombre}")
    if len(ccaas_db) > 5:
        print(f"  ... y {len(ccaas_db)-5} más")
    
    # Resumen por tipo de jornada
    print("\nResumen por tipo de jornada:")
    for jornada in df['tipo_jornada'].unique():
        count = len(df[df['tipo_jornada'] == jornada])
        print(f"  - {jornada}: {count} registros")
    
    # Resumen por nivel CNAE
    print("\nResumen por nivel CNAE:")
    for nivel in df['cnae_nivel'].unique():
        count = len(df[df['cnae_nivel'] == nivel])
        print(f"  - {nivel}: {count} registros")
    
    # Resumen por métrica
    print("\nResumen por métrica:")
    for metrica in df['metrica'].unique():
        count = len(df[df['metrica'] == metrica])
        print(f"  - {metrica}: {count} registros")
    
    return df

def create_detailed_comparison(df_ine, df_db, ccaa_col, sector_col, jornada_col):
    """
    Crea comparación detallada entre INE y DuckDB para tabla 6063
    """
    print("\n" + "=" * 80)
    print("REALIZANDO COMPARACIÓN DETALLADA")
    print("=" * 80)
    
    comparisons = []
    
    # Mapeos necesarios
    jornada_map = {
        'Ambas jornadas': 'TOTAL',
        'Jornada a tiempo completo': 'COMPLETA',
        'Jornada a tiempo parcial': 'PARCIAL',
        'Total': 'TOTAL'
    }
    
    sector_map = {
        'B_S Industria, construcción y servicios (excepto actividades de los hogares como empleadores y de organizaciones y organismos extraterritoriales)': ('TOTAL', None),
        'Total': ('TOTAL', None),
        'Industria': ('SECTOR_BS', 'B-E'),
        'Construcción': ('SECTOR_BS', 'F'),
        'Servicios': ('SECTOR_BS', 'G-S')
    }
    
    # Mapeo de CCAA - basado en formato "01 Andalucía"
    ccaa_map = {
        'Total Nacional': ('NAC', None),
        '01 Andalucía': ('CCAA', '01'),
        '02 Aragón': ('CCAA', '02'),
        '03 Asturias, Principado de': ('CCAA', '03'),
        '04 Balears, Illes': ('CCAA', '04'),
        '05 Canarias': ('CCAA', '05'),
        '06 Cantabria': ('CCAA', '06'),
        '07 Castilla y León': ('CCAA', '07'),
        '08 Castilla - La Mancha': ('CCAA', '08'),
        '09 Cataluña': ('CCAA', '09'),
        '10 Comunitat Valenciana': ('CCAA', '10'),
        '11 Extremadura': ('CCAA', '11'),
        '12 Galicia': ('CCAA', '12'),
        '13 Madrid, Comunidad de': ('CCAA', '13'),
        '14 Murcia, Región de': ('CCAA', '14'),
        '15 Navarra, Comunidad Foral de': ('CCAA', '15'),
        '16 País Vasco': ('CCAA', '16'),
        '17 Rioja, La': ('CCAA', '17')
    }
    
    metrica_map = {
        'Horas pactadas': 'horas_pactadas',
        'Horas pagadas': 'horas_pagadas',
        'Horas efectivas': 'horas_efectivas',
        'Horas efectivas de trabajo': 'horas_efectivas',
        'Horas extraordinarias': 'horas_extraordinarias',
        'Horas extras por trabajador': 'horas_extraordinarias',
        'Horas no trabajadas': 'horas_no_trabajadas'
    }
    
    # Procesar cada registro del INE
    total_ine = len(df_ine)
    procesados = 0
    no_encontrados = []
    
    for idx, row_ine in df_ine.iterrows():
        procesados += 1
        if procesados % 100 == 0:
            print(f"  Procesando: {procesados}/{total_ine}")
        
        ccaa_ine = row_ine[ccaa_col] if ccaa_col else ''
        sector_ine = row_ine[sector_col] if sector_col else ''
        jornada_ine = row_ine[jornada_col] if jornada_col else ''
        tiempo_ine = row_ine['Tiempo de trabajo']
        valor_ine = row_ine['Valor']
        
        # Mapear valores
        jornada_db = jornada_map.get(jornada_ine, jornada_ine)
        
        # Mapear CCAA
        ambito_territorial = None
        ccaa_codigo = None
        if ccaa_ine in ccaa_map:
            ambito_territorial, ccaa_codigo = ccaa_map[ccaa_ine]
        else:
            # Si no está en el mapeo, intentar identificar
            if 'Total' in ccaa_ine or 'Nacional' in ccaa_ine:
                ambito_territorial, ccaa_codigo = 'NAC', None
            else:
                # Intentar extraer código de CCAA
                import re
                match = re.match(r'^(\d{2})\s', ccaa_ine)
                if match:
                    ambito_territorial = 'CCAA'
                    ccaa_codigo = match.group(1)
        
        # Mapear sector
        if sector_ine in sector_map:
            cnae_nivel, cnae_codigo = sector_map[sector_ine]
        else:
            if 'B_S' in sector_ine or 'Total' in sector_ine:
                cnae_nivel, cnae_codigo = 'TOTAL', None
            else:
                cnae_nivel, cnae_codigo = None, None
        
        metrica_db = metrica_map.get(tiempo_ine, None)
        
        if metrica_db and cnae_nivel and ambito_territorial:
            # Buscar valor correspondiente en DuckDB
            mask = (
                (df_db['tipo_jornada'] == jornada_db) &
                (df_db['cnae_nivel'] == cnae_nivel) &
                (df_db['metrica'] == metrica_db) &
                (df_db['ambito_territorial'] == ambito_territorial)
            )
            
            if cnae_codigo:
                mask = mask & (df_db['cnae_codigo'] == cnae_codigo)
            else:
                mask = mask & (df_db['cnae_codigo'].isna())
            
            if ccaa_codigo:
                mask = mask & (df_db['ccaa_codigo'] == ccaa_codigo)
            else:
                mask = mask & (df_db['ccaa_codigo'].isna())
            
            # Para HNT, solo tomar el total (causa = NULL)
            if metrica_db == 'horas_no_trabajadas':
                mask = mask & (df_db['causa'].isna())
            
            df_match = df_db[mask]
            
            if len(df_match) > 0:
                valor_db = df_match.iloc[0]['valor']
                diferencia = abs(valor_ine - valor_db)
                pct_diff = (diferencia / valor_ine * 100) if valor_ine != 0 else 0
                
                if diferencia < 0.1:
                    estado = 'OK'
                elif diferencia < 1:
                    estado = 'ADVERTENCIA'
                else:
                    estado = 'ERROR'
                
                comparisons.append({
                    'CCAA': ccaa_ine[:30],
                    'Sector': sector_ine[:30],
                    'Jornada': jornada_ine,
                    'Métrica': tiempo_ine,
                    'Valor INE': valor_ine,
                    'Valor DuckDB': valor_db,
                    'Diferencia': round(diferencia, 2),
                    '% Diff': round(pct_diff, 2),
                    'Estado': estado
                })
            else:
                no_encontrados.append((ccaa_ine, sector_ine, jornada_ine, tiempo_ine))
                comparisons.append({
                    'CCAA': ccaa_ine[:30],
                    'Sector': sector_ine[:30],
                    'Jornada': jornada_ine,
                    'Métrica': tiempo_ine,
                    'Valor INE': valor_ine,
                    'Valor DuckDB': 'NO ENCONTRADO',
                    'Diferencia': '-',
                    '% Diff': '-',
                    'Estado': 'SIN DATOS'
                })
    
    df_comparison = pd.DataFrame(comparisons)
    
    # Mostrar resumen
    print(f"\nTotal comparaciones: {len(comparisons)}")
    print("\nResumen por estado:")
    for estado in df_comparison['Estado'].unique():
        count = len(df_comparison[df_comparison['Estado'] == estado])
        print(f"  - {estado}: {count}")
    
    if no_encontrados:
        print(f"\nRegistros no encontrados: {len(no_encontrados)}")
        for i, (ccaa, sector, jornada, metrica) in enumerate(no_encontrados[:3]):
            print(f"  - {ccaa} / {sector[:20]} / {jornada} / {metrica}")
        if len(no_encontrados) > 3:
            print(f"  ... y {len(no_encontrados)-3} más")
    
    return df_comparison

def create_validation_excel(df_ine, df_db, df_comparison, output_path):
    """
    Crea Excel detallado con todas las validaciones
    """
    print("\n" + "=" * 80)
    print("GENERANDO EXCEL DE VALIDACIÓN")
    print("=" * 80)
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    info_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Hoja 1: Comparación Detallada
    ws1 = wb.active
    ws1.title = "Comparacion"
    
    # Escribir datos de comparación
    for r_idx, row in enumerate(dataframe_to_rows(df_comparison, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:  # Headers
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Colorear según estado
                estado = ws1.cell(row=r_idx, column=9).value  # Columna Estado
                if estado == 'OK':
                    for col in range(1, 10):
                        ws1.cell(row=r_idx, column=col).fill = ok_fill
                elif estado == 'ERROR':
                    for col in range(1, 10):
                        ws1.cell(row=r_idx, column=col).fill = error_fill
                elif estado == 'ADVERTENCIA':
                    for col in range(1, 10):
                        ws1.cell(row=r_idx, column=col).fill = warning_fill
                elif estado == 'SIN DATOS':
                    for col in range(1, 10):
                        ws1.cell(row=r_idx, column=col).fill = info_fill
    
    # Ajustar anchos de columna
    column_widths = [30, 30, 20, 25, 12, 12, 12, 10, 12]
    for i, width in enumerate(column_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = width
    
    # Hoja 2: Resumen
    ws2 = wb.create_sheet("Resumen")
    
    # Título
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "VALIDACIÓN TABLA 6063 - PERIODO 2025T1"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    ws2['A3'] = "CCAA + Sectores B-S + Tipo de Jornada"
    ws2['A3'].font = Font(italic=True)
    
    # Estadísticas
    ws2['A5'] = "ESTADÍSTICAS GENERALES"
    ws2['A5'].font = Font(bold=True, size=12)
    
    stats = df_comparison['Estado'].value_counts()
    row = 6
    for estado, count in stats.items():
        ws2.cell(row=row, column=1, value=f"{estado}:")
        ws2.cell(row=row, column=2, value=count)
        
        # Colorear según estado
        if estado == 'OK':
            ws2.cell(row=row, column=2).fill = ok_fill
        elif estado == 'ERROR':
            ws2.cell(row=row, column=2).fill = error_fill
        elif estado == 'ADVERTENCIA':
            ws2.cell(row=row, column=2).fill = warning_fill
        
        row += 1
    
    row += 1
    ws2.cell(row=row, column=1, value="TOTAL:")
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=len(df_comparison))
    ws2.cell(row=row, column=2).font = Font(bold=True)
    
    # Porcentaje de éxito
    ok_count = len(df_comparison[df_comparison['Estado'] == 'OK'])
    success_rate = (ok_count / len(df_comparison) * 100) if len(df_comparison) > 0 else 0
    
    row += 2
    ws2.cell(row=row, column=1, value="Tasa de éxito:")
    ws2.cell(row=row, column=2, value=f"{success_rate:.1f}%")
    ws2.cell(row=row, column=2).font = Font(bold=True, size=12)
    
    if success_rate >= 95:
        ws2.cell(row=row, column=2).fill = ok_fill
    elif success_rate >= 80:
        ws2.cell(row=row, column=2).fill = warning_fill
    else:
        ws2.cell(row=row, column=2).fill = error_fill
    
    # Información adicional
    row += 2
    ws2.cell(row=row, column=1, value="DETALLES:")
    ws2.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws2.cell(row=row, column=1, value="Registros INE:")
    ws2.cell(row=row, column=2, value=len(df_ine))
    row += 1
    ws2.cell(row=row, column=1, value="Registros DuckDB:")
    ws2.cell(row=row, column=2, value=len(df_db))
    row += 1
    ws2.cell(row=row, column=1, value="Dimensiones:")
    ws2.cell(row=row, column=2, value="CCAA + Sector + Jornada")
    
    # Hoja 3: Datos INE Original
    ws3 = wb.create_sheet("INE_Original")
    for r_idx, row in enumerate(dataframe_to_rows(df_ine, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    # Hoja 4: Datos DuckDB
    ws4 = wb.create_sheet("DuckDB_Datos")
    for r_idx, row in enumerate(dataframe_to_rows(df_db, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    # Guardar
    wb.save(output_path)
    print(f"Excel guardado: {output_path}")
    
    return output_path

def main():
    """
    Función principal de validación tabla 6063
    """
    print("\n" + "=" * 80)
    print("VALIDACIÓN DETALLADA TABLA 6063")
    print("CCAA + Sectores B-S + Tipo de Jornada")
    print("Comparación INE CSV vs DuckDB - Periodo 2025T1")
    print("=" * 80 + "\n")
    
    # Configuración de rutas
    base_dir = Path(__file__).parent
    csv_file = base_dir / "data" / "INE" / "6063.csv"
    db_path = base_dir / "data" / "analysis.db"
    output_file = base_dir / "data" / "INE" / "validacion_6063_detallada.xlsx"
    
    try:
        # 1. Leer y preparar datos del INE
        df_ine, ccaa_col, sector_col, jornada_col = read_and_prepare_ine_csv(csv_file)
        
        # 2. Obtener datos de DuckDB
        df_db = get_duckdb_data_detailed(db_path)
        
        # 3. Realizar comparación detallada
        df_comparison = create_detailed_comparison(df_ine, df_db, ccaa_col, sector_col, jornada_col)
        
        # 4. Crear Excel de validación
        output_path = create_validation_excel(df_ine, df_db, df_comparison, output_file)
        
        # 5. Mostrar resumen final
        print("\n" + "=" * 80)
        print("VALIDACIÓN COMPLETADA - TABLA 6063")
        print("=" * 80)
        
        # Calcular estadísticas finales
        ok_count = len(df_comparison[df_comparison['Estado'] == 'OK'])
        total_count = len(df_comparison)
        success_rate = (ok_count / total_count * 100) if total_count > 0 else 0
        
        print(f"\nResultados:")
        print(f"  - Comparaciones exitosas: {ok_count}/{total_count}")
        print(f"  - Tasa de éxito: {success_rate:.1f}%")
        
        if success_rate >= 95:
            print(f"\n[OK] Validación exitosa - Alta coincidencia con datos INE")
        elif success_rate >= 80:
            print(f"\n[ADVERTENCIA] Validación con observaciones - Revisar discrepancias")
        else:
            print(f"\n[ERROR] Validación con problemas - Requiere revisión detallada")
        
        print(f"\nArchivo Excel generado: {output_file.name}")
        print(f"Abre el archivo para revisar:")
        print(f"  1. Hoja 'Comparacion': Detalle línea por línea")
        print(f"  2. Hoja 'Resumen': Estadísticas de validación")
        print(f"  3. Hoja 'INE_Original': Datos fuente del INE")
        print(f"  4. Hoja 'DuckDB_Datos': Datos en nuestra BD")
        
        print(f"\nCaracterística especial tabla 6063:")
        print(f"  - 17 CCAA + Total Nacional")
        print(f"  - Sectores B-S agregados")
        print(f"  - CON tipo de jornada (3 tipos)")
        
    except Exception as e:
        print(f"\nError durante la validación: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()