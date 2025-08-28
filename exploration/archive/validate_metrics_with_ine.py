#!/usr/bin/env python3
"""
Script para validar las métricas extraídas contra la metodología oficial del INE.
Compara las métricas encontradas con las definidas en el documento oficial.
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configuración de rutas
BASE_DIR = Path(__file__).parent.parent
REPORTS_DIR = BASE_DIR / "data" / "exploration_reports"

# MÉTRICAS OFICIALES DEL INE - Sección 9 del PDF de Metodología
METRICAS_OFICIALES_INE = {
    'COSTES_LABORALES': [
        'Coste total',
        'Coste laboral total',
        'Coste salarial total',
        'Coste salarial ordinario',
        'Coste salarial pagos extraordinarios',
        'Coste salarial pagos atrasados',
        'Coste salarial extraordinario',
        'Otros costes',
        'Coste por percepciones no salariales',
        'Coste por I.T.',
        'Coste por desempleo',
        'Coste por otras prestaciones sociales directas',
        'Coste por otras percepciones no salariales',
        'Coste por cotizaciones obligatorias',
        'Coste por contingencias comunes',
        'Coste por desempleo, Fogasa y F. Profesional',
        'Coste por otras cotizaciones sociales obligatorias',
        'Subvenciones y bonificaciones de la Seguridad Social',
        'Coste por despido',
        'Percepciones por día de I.T.',
        'Coste indemnización trabajador despedido',
        'Coste por hora extra'
    ],
    'TIEMPO_TRABAJO': [
        'Horas pactadas',
        'Horas pagadas',
        'Horas efectivas',
        'Horas no trabajadas',
        'Horas no trabajadas por vacaciones',
        'Horas no trabajadas por fiestas',
        'Horas no trabajadas por I.T.',
        'Horas no trabajadas por maternidad',
        'Horas no trabajadas por permisos remunerados',
        'Horas no trabajadas por razones técnicas',
        'Horas no trabajadas por conflictos laborales',
        'Horas extras',
        'Horas extraordinarias'
    ],
    'VACANTES': [
        'Número de vacantes',
        'Motivos por los que no existen vacantes'
    ]
}

# Cargar nombres alternativos y variaciones comunes
EQUIVALENCIAS = {
    # Costes
    'Coste laboral total': ['Coste total por trabajador', 'Coste laboral'],
    'Coste por I.T.': ['Coste I.T.', 'Pagos por Incapacidad Temporal', 'Coste por incapacidad temporal'],
    'Coste por desempleo': ['Coste por desempleo parcial', 'Pagos por desempleo'],
    'Coste por despido': ['Indemnizaciones por despido', 'Coste indemnización'],
    'Coste por contingencias comunes': ['Contingencias comunes'],
    'Coste por desempleo, Fogasa y F. Profesional': ['Desempleo, Fogasa y Formación Profesional'],
    'Subvenciones y bonificaciones': ['Subvenciones y bonificaciones de la S.Social'],
    
    # Tiempo
    'Horas efectivas': ['Horas realmente trabajadas'],
    'Horas no trabajadas': ['Tiempo no trabajado'],
    'Horas extras': ['Horas extraordinarias', 'Horas extras por trabajador'],
    'Horas no trabajadas por I.T.': ['Horas no trabajadas por I.T', 'Horas no trabajadas por incapacidad temporal'],
    'Horas no trabajadas por vacaciones': ['Horas no trabajadas por vacaciones y fiestas'],
    
    # Vacantes
    'Número de vacantes': ['Vacantes', 'Puestos vacantes'],
    'Motivos por los que no existen vacantes': ['Motivos no vacantes', 'Razones de no vacantes']
}

def normalize_metric_name(name: str) -> str:
    """Normaliza un nombre de métrica para comparación."""
    # Convertir a minúsculas y eliminar espacios extra
    normalized = name.lower().strip()
    # Eliminar caracteres especiales comunes
    normalized = normalized.replace('.', '').replace(',', '')
    # Normalizar espacios
    normalized = ' '.join(normalized.split())
    return normalized

def find_metric_in_official(metric: str, official_metrics: Dict[str, List[str]]) -> tuple:
    """
    Busca una métrica en el listado oficial.
    Retorna (categoría, métrica_oficial, coincidencia_exacta)
    """
    metric_norm = normalize_metric_name(metric)
    
    # Buscar coincidencia exacta o equivalente
    for category, metrics_list in official_metrics.items():
        for official_metric in metrics_list:
            official_norm = normalize_metric_name(official_metric)
            
            # Coincidencia exacta
            if metric_norm == official_norm:
                return category, official_metric, True
            
            # Buscar en equivalencias
            if official_metric in EQUIVALENCIAS:
                for equiv in EQUIVALENCIAS[official_metric]:
                    if metric_norm == normalize_metric_name(equiv):
                        return category, official_metric, True
            
            # Coincidencia parcial (una contenida en la otra)
            if metric_norm in official_norm or official_norm in metric_norm:
                return category, official_metric, False
    
    return None, None, False

def load_extracted_metrics(json_file: Path) -> Dict:
    """Carga las métricas extraídas del archivo JSON más reciente."""
    if not json_file.exists():
        # Buscar el archivo más reciente
        json_files = list(REPORTS_DIR.glob("metricas_limpias_*.json"))
        if not json_files:
            raise FileNotFoundError("No se encontró archivo de métricas extraídas")
        json_file = max(json_files, key=lambda x: x.stat().st_mtime)
    
    with open(json_file, 'r', encoding='utf-8') as f:
        return json.load(f)

def validate_metrics(extracted_data: Dict) -> Dict:
    """Valida las métricas extraídas contra las oficiales del INE."""
    
    # Recopilar todas las métricas únicas extraídas
    extracted_metrics = set()
    for table_data in extracted_data.values():
        if table_data and 'metricas' in table_data:
            for metric in table_data['metricas']:
                extracted_metrics.add(metric['nombre'])
    
    # Validar cada métrica extraída
    validation_results = {
        'metricas_validadas': [],
        'metricas_no_encontradas': [],
        'metricas_parciales': [],
        'resumen_por_categoria': {}
    }
    
    for metric in sorted(extracted_metrics):
        category, official, exact = find_metric_in_official(metric, METRICAS_OFICIALES_INE)
        
        if category and exact:
            validation_results['metricas_validadas'].append({
                'metrica_extraida': metric,
                'metrica_oficial': official,
                'categoria': category
            })
        elif category and not exact:
            validation_results['metricas_parciales'].append({
                'metrica_extraida': metric,
                'posible_oficial': official,
                'categoria': category
            })
        else:
            validation_results['metricas_no_encontradas'].append(metric)
    
    # Buscar métricas oficiales no extraídas
    all_official = []
    for cat, metrics in METRICAS_OFICIALES_INE.items():
        all_official.extend([(m, cat) for m in metrics])
    
    extracted_normalized = {normalize_metric_name(m) for m in extracted_metrics}
    
    validation_results['metricas_oficiales_faltantes'] = []
    for official_metric, category in all_official:
        found = False
        official_norm = normalize_metric_name(official_metric)
        
        # Verificar si está extraída
        if official_norm in extracted_normalized:
            found = True
        else:
            # Verificar equivalencias
            if official_metric in EQUIVALENCIAS:
                for equiv in EQUIVALENCIAS[official_metric]:
                    if normalize_metric_name(equiv) in extracted_normalized:
                        found = True
                        break
        
        if not found:
            # Verificar coincidencias parciales
            for extracted in extracted_normalized:
                if official_norm in extracted or extracted in official_norm:
                    found = True
                    break
        
        if not found:
            validation_results['metricas_oficiales_faltantes'].append({
                'metrica': official_metric,
                'categoria': category
            })
    
    # Calcular resumen
    for cat in METRICAS_OFICIALES_INE.keys():
        total_oficiales = len(METRICAS_OFICIALES_INE[cat])
        encontradas = len([m for m in validation_results['metricas_validadas'] 
                          if m['categoria'] == cat])
        parciales = len([m for m in validation_results['metricas_parciales'] 
                        if m['categoria'] == cat])
        
        validation_results['resumen_por_categoria'][cat] = {
            'total_oficiales': total_oficiales,
            'encontradas_exactas': encontradas,
            'encontradas_parciales': parciales,
            'porcentaje_cobertura': round((encontradas / total_oficiales) * 100, 1) if total_oficiales > 0 else 0
        }
    
    # Resumen global
    total_oficiales = sum(len(m) for m in METRICAS_OFICIALES_INE.values())
    total_validadas = len(validation_results['metricas_validadas'])
    total_parciales = len(validation_results['metricas_parciales'])
    
    validation_results['resumen_global'] = {
        'total_metricas_oficiales': total_oficiales,
        'total_metricas_extraidas': len(extracted_metrics),
        'metricas_validadas': total_validadas,
        'metricas_parciales': total_parciales,
        'metricas_no_encontradas': len(validation_results['metricas_no_encontradas']),
        'porcentaje_precision': round((total_validadas / len(extracted_metrics)) * 100, 1) if extracted_metrics else 0,
        'porcentaje_cobertura': round((total_validadas / total_oficiales) * 100, 1) if total_oficiales > 0 else 0
    }
    
    return validation_results

def generate_validation_report(validation: Dict, output_file: Path):
    """Genera un informe Excel de validación."""
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Hoja 1: Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Validación"
    
    # Título
    ws_resumen.merge_cells('A1:D1')
    ws_resumen['A1'] = "VALIDACIÓN DE MÉTRICAS CONTRA METODOLOGÍA OFICIAL INE"
    ws_resumen['A1'].font = Font(size=14, bold=True)
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    # Resumen global
    row = 3
    ws_resumen[f'A{row}'] = "RESUMEN GLOBAL"
    ws_resumen[f'A{row}'].font = Font(bold=True)
    row += 1
    
    global_data = validation['resumen_global']
    ws_resumen[f'A{row}'] = "Total métricas oficiales INE:"
    ws_resumen[f'B{row}'] = global_data['total_metricas_oficiales']
    row += 1
    ws_resumen[f'A{row}'] = "Total métricas extraídas:"
    ws_resumen[f'B{row}'] = global_data['total_metricas_extraidas']
    row += 1
    ws_resumen[f'A{row}'] = "Métricas validadas (exactas):"
    ws_resumen[f'B{row}'] = global_data['metricas_validadas']
    ws_resumen[f'B{row}'].fill = success_fill
    row += 1
    ws_resumen[f'A{row}'] = "Métricas parciales:"
    ws_resumen[f'B{row}'] = global_data['metricas_parciales']
    ws_resumen[f'B{row}'].fill = warning_fill
    row += 1
    ws_resumen[f'A{row}'] = "Métricas no encontradas:"
    ws_resumen[f'B{row}'] = global_data['metricas_no_encontradas']
    ws_resumen[f'B{row}'].fill = error_fill if global_data['metricas_no_encontradas'] > 0 else success_fill
    row += 2
    
    ws_resumen[f'A{row}'] = "% Precisión (validadas/extraídas):"
    ws_resumen[f'B{row}'] = f"{global_data['porcentaje_precision']}%"
    row += 1
    ws_resumen[f'A{row}'] = "% Cobertura (validadas/oficiales):"
    ws_resumen[f'B{row}'] = f"{global_data['porcentaje_cobertura']}%"
    
    # Resumen por categoría
    row += 3
    ws_resumen[f'A{row}'] = "RESUMEN POR CATEGORÍA"
    ws_resumen[f'A{row}'].font = Font(bold=True)
    row += 1
    
    headers = ["Categoría", "Total Oficial", "Encontradas", "Parciales", "% Cobertura"]
    for col, header in enumerate(headers, 1):
        cell = ws_resumen.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    for cat, data in validation['resumen_por_categoria'].items():
        ws_resumen.cell(row=row, column=1, value=cat)
        ws_resumen.cell(row=row, column=2, value=data['total_oficiales'])
        ws_resumen.cell(row=row, column=3, value=data['encontradas_exactas'])
        ws_resumen.cell(row=row, column=4, value=data['encontradas_parciales'])
        ws_resumen.cell(row=row, column=5, value=f"{data['porcentaje_cobertura']}%")
        
        # Colorear según cobertura
        if data['porcentaje_cobertura'] >= 80:
            ws_resumen.cell(row=row, column=5).fill = success_fill
        elif data['porcentaje_cobertura'] >= 50:
            ws_resumen.cell(row=row, column=5).fill = warning_fill
        else:
            ws_resumen.cell(row=row, column=5).fill = error_fill
        row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 20
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 15
    ws_resumen.column_dimensions['E'].width = 15
    
    # Hoja 2: Métricas Validadas
    ws_validadas = wb.create_sheet("Métricas Validadas")
    headers = ["Métrica Extraída", "Métrica Oficial INE", "Categoría"]
    for col, header in enumerate(headers, 1):
        cell = ws_validadas.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 2
    for item in validation['metricas_validadas']:
        ws_validadas.cell(row=row, column=1, value=item['metrica_extraida'])
        ws_validadas.cell(row=row, column=2, value=item['metrica_oficial'])
        ws_validadas.cell(row=row, column=3, value=item['categoria'])
        row += 1
    
    ws_validadas.column_dimensions['A'].width = 50
    ws_validadas.column_dimensions['B'].width = 50
    ws_validadas.column_dimensions['C'].width = 25
    
    # Hoja 3: Métricas Oficiales Faltantes
    ws_faltantes = wb.create_sheet("Métricas INE Faltantes")
    headers = ["Métrica Oficial INE", "Categoría"]
    for col, header in enumerate(headers, 1):
        cell = ws_faltantes.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 2
    for item in validation['metricas_oficiales_faltantes']:
        ws_faltantes.cell(row=row, column=1, value=item['metrica'])
        ws_faltantes.cell(row=row, column=2, value=item['categoria'])
        ws_faltantes.cell(row=row, column=1).fill = warning_fill
        row += 1
    
    ws_faltantes.column_dimensions['A'].width = 50
    ws_faltantes.column_dimensions['B'].width = 25
    
    # Hoja 4: Métricas No Encontradas
    if validation['metricas_no_encontradas']:
        ws_no_encontradas = wb.create_sheet("Métricas No Validadas")
        ws_no_encontradas.cell(row=1, column=1, value="Métrica Extraída (No encontrada en metodología INE)")
        ws_no_encontradas.cell(row=1, column=1).fill = header_fill
        ws_no_encontradas.cell(row=1, column=1).font = header_font
        
        row = 2
        for metric in validation['metricas_no_encontradas']:
            ws_no_encontradas.cell(row=row, column=1, value=metric)
            ws_no_encontradas.cell(row=row, column=1).fill = error_fill
            row += 1
        
        ws_no_encontradas.column_dimensions['A'].width = 60
    
    # Guardar
    wb.save(output_file)
    print(f"\n[OK] Informe de validación guardado en: {output_file}")

def main():
    """Función principal."""
    print("=" * 80)
    print("VALIDACIÓN DE MÉTRICAS CONTRA METODOLOGÍA OFICIAL INE")
    print("=" * 80)
    
    try:
        # Buscar archivo de métricas más reciente (ahora buscamos enhanced)
        json_files = list(REPORTS_DIR.glob("metricas_enhanced_*.json"))
        if not json_files:
            # Si no hay enhanced, buscar limpias
            json_files = list(REPORTS_DIR.glob("metricas_limpias_*.json"))
            if not json_files:
                print("[ERROR] No se encontró archivo de métricas extraídas")
                print("Ejecute primero: python exploration/extract_metrics_enhanced.py")
                return
        
        latest_json = max(json_files, key=lambda x: x.stat().st_mtime)
        print(f"\nCargando métricas extraídas de: {latest_json.name}")
        
        # Cargar datos
        extracted_data = load_extracted_metrics(latest_json)
        
        # Validar
        print("\nValidando contra metodología oficial del INE...")
        validation_results = validate_metrics(extracted_data)
        
        # Generar timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Guardar resultados JSON
        json_output = REPORTS_DIR / f"validacion_ine_{timestamp}.json"
        with open(json_output, 'w', encoding='utf-8') as f:
            json.dump(validation_results, f, ensure_ascii=False, indent=2)
        print(f"\n[OK] Resultados JSON guardados en: {json_output}")
        
        # Generar Excel
        excel_output = REPORTS_DIR / f"validacion_ine_{timestamp}.xlsx"
        generate_validation_report(validation_results, excel_output)
        
        # Mostrar resumen
        print("\n" + "=" * 80)
        print("RESUMEN DE VALIDACIÓN")
        print("=" * 80)
        
        global_summary = validation_results['resumen_global']
        print(f"\nMétricas oficiales INE: {global_summary['total_metricas_oficiales']}")
        print(f"Métricas extraídas: {global_summary['total_metricas_extraidas']}")
        print(f"Métricas validadas: {global_summary['metricas_validadas']} ({global_summary['porcentaje_precision']:.1f}% precisión)")
        print(f"Cobertura oficial: {global_summary['porcentaje_cobertura']:.1f}%")
        
        print("\nPor categoría:")
        for cat, data in validation_results['resumen_por_categoria'].items():
            print(f"  {cat}: {data['encontradas_exactas']}/{data['total_oficiales']} ({data['porcentaje_cobertura']}%)")
        
        if validation_results['metricas_oficiales_faltantes']:
            print(f"\n[!] Métricas oficiales no detectadas: {len(validation_results['metricas_oficiales_faltantes'])}")
            print("  Ver hoja 'Métricas INE Faltantes' en el Excel para detalles")
        
        print("\n[OK] Validación completada")
        print(f"[OK] Revise el archivo Excel para el informe detallado")
        
    except Exception as e:
        print(f"\n[ERROR] {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()