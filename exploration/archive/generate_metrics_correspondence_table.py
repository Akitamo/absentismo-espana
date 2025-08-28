#!/usr/bin/env python3
"""
Script para generar tabla de correspondencia entre métricas del documento INE 
y las métricas extraídas, indicando de qué tablas se obtiene cada una.
"""

import json
from pathlib import Path
from datetime import datetime
import pandas as pd
from typing import Dict, List, Set, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de rutas
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data" / "raw" / "csv"
REPORTS_DIR = BASE_DIR / "data" / "exploration_reports"

# MÉTRICAS OFICIALES DEL INE según documento metodología_ETCL_INE_2023.pdf
METRICAS_INE = {
    'COSTES_LABORALES': [
        ('Coste total', 'Coste laboral total más Otros costes'),
        ('Otros costes', 'Percepciones no salariales más cotizaciones obligatorias menos subvenciones'),
        ('Coste salarial total', 'Incluye costes salariales ordinarios, pagos extraordinarios y pagos atrasados'),
        ('Coste salarial ordinario', 'Pagos salariales de periodicidad mensual'),
        ('Coste salarial pagos extraordinarios', 'Pagas extraordinarias y pagos de vencimiento superior al mes'),
        ('Coste salarial pagos atrasados', 'Pagos efectuados en el mes y devengados en periodos anteriores'),
        ('Coste salarial extraordinario', 'Coste salarial pagos extraordinarios más pagos atrasados'),
        ('Coste por percepciones no salariales', 'Incluye I.T., desempleo, prestaciones directas, indemnizaciones'),
        ('Coste I.T.', 'Pagos por I.T. a cargo del empleador'),
        ('Coste por desempleo', 'Pagos por desempleo (reducción jornada/suspensión) a cargo empleador'),
        ('Coste por otras prestaciones sociales directas', 'Prestaciones complementarias a la Seguridad Social'),
        ('Coste por otras percepciones no salariales', 'Resto de percepciones no salariales'),
        ('Coste por cotizaciones obligatorias', 'Coste de la Seguridad Social obligatoria del empleador'),
        ('Coste por Contingencias Comunes', 'Cotizaciones por Contingencias Comunes de la S.S.'),
        ('Coste por Desempleo, Fogasa y F. Profesional', 'Cotizaciones que cubren estas contingencias'),
        ('Coste por otras cotizaciones sociales obligatorias', 'Resto de cotizaciones obligatorias S.S.'),
        ('Subvenciones y bonificaciones de la S.Social', 'Reducciones y bonificaciones en liquidaciones S.S.'),
        ('Coste por despido', 'Pagos por indemnización por despido y extinción de contrato'),
        ('Percepciones por día de I.T.', 'Pagos de I.T por día que el trabajador causa baja'),
        ('Coste indemnización trabajador despedido', 'Cociente del coste despido entre trabajadores despedidos'),
        ('Coste por hora extra', 'Coste de horas extras incluyendo pagos, cotizaciones y compensaciones')
    ],
    'TIEMPO_TRABAJO': [
        ('Horas pactadas', 'Horas legalmente establecidas por acuerdo empleador/trabajadores'),
        ('Horas pagadas', 'Comprende horas trabajadas y no trabajadas remuneradas'),
        ('Horas efectivas', 'Horas realmente trabajadas incluyendo extraordinarias'),
        ('Horas no trabajadas', 'Total de horas pactadas no trabajadas por cualquier motivo'),
        ('Horas no trabajadas por vacaciones', 'No trabajadas por vacaciones'),
        ('Horas no trabajadas por fiestas', 'No trabajadas por fiestas oficiales o no oficiales'),
        ('Horas no trabajadas por I.T.', 'No trabajadas por incapacidad temporal'),
        ('Horas no trabajadas por maternidad', 'No trabajadas por maternidad, adopción'),
        ('Horas no trabajadas por permisos remunerados', 'Permisos por nupcialidad, natalidad, fallecimiento'),
        ('Horas no trabajadas por razones técnicas', 'Por razones técnicas/económicas con o sin ERE'),
        ('Horas no trabajadas por conflictos laborales', 'Por huelgas y conflictos laborales'),
        ('Horas no trabajadas (otras)', 'Representación sindical, visitas médicas, fuerza mayor, absentismo'),
        ('Horas extras', 'Horas extraordinarias realizadas')
    ],
    'COSTE_SALARIAL': [
        ('Salario Base', 'Retribución fija del trabajador'),
        ('Complementos Salariales', 'Complementos al salario base'),
        ('Pagos por Horas Extraordinarias', 'Pagos por horas extras estructurales y no estructurales'),
        ('Gratificaciones Extraordinarias', 'Retribuciones con periodicidad superior al mes')
    ],
    'VACANTES': [
        ('Número de vacantes', 'Número de vacantes en el trimestre de referencia'),
        ('Motivos por los que no existen vacantes', 'Distribución porcentual de motivos de no vacantes')
    ]
}

def load_extracted_metrics() -> Dict:
    """Carga las métricas extraídas del archivo JSON más reciente."""
    json_files = list(REPORTS_DIR.glob("metricas_limpias_*.json"))
    if not json_files:
        raise FileNotFoundError("No se encontró archivo de métricas extraídas")
    
    latest_file = max(json_files, key=lambda x: x.stat().st_mtime)
    print(f"Cargando métricas de: {latest_file.name}")
    
    with open(latest_file, 'r', encoding='utf-8') as f:
        return json.load(f)

def normalize_metric_name(name: str) -> str:
    """Normaliza nombres de métricas para comparación."""
    normalized = name.lower().strip()
    # Eliminar caracteres especiales pero mantener puntos para I.T.
    normalized = normalized.replace(',', '').replace(':', '')
    # Normalizar espacios
    normalized = ' '.join(normalized.split())
    # Normalizar variaciones comunes
    normalized = normalized.replace('i.t.', 'it').replace('i.t', 'it')
    normalized = normalized.replace('fogasa', 'fogasa')
    normalized = normalized.replace('f. profesional', 'formación profesional')
    return normalized

def find_metric_correspondence(ine_metric: str, extracted_data: Dict) -> List[Tuple[str, str]]:
    """
    Encuentra en qué tablas aparece una métrica del INE.
    Retorna lista de tuplas (código_tabla, nombre_métrica_extraída)
    """
    ine_norm = normalize_metric_name(ine_metric)
    tables_with_metric = []
    
    for table_code, table_data in extracted_data.items():
        if not table_data or 'metricas' not in table_data:
            continue
            
        for metric in table_data['metricas']:
            metric_norm = normalize_metric_name(metric['nombre'])
            
            # Coincidencia exacta o muy similar
            if (ine_norm == metric_norm or 
                ine_norm in metric_norm or 
                metric_norm in ine_norm or
                # Casos especiales
                (ine_norm == 'coste total' and metric_norm == 'coste laboral total') or
                (ine_norm == 'horas extras' and 'horas extra' in metric_norm) or
                (ine_norm == 'coste it' and 'coste por it' in metric_norm) or
                (ine_norm == 'horas no trabajadas por vacaciones' and 'vacaciones y fiestas' in metric_norm)):
                
                tables_with_metric.append((table_code, metric['nombre']))
                break
    
    return tables_with_metric

def generate_correspondence_table(extracted_data: Dict) -> pd.DataFrame:
    """Genera DataFrame con la correspondencia de métricas."""
    rows = []
    
    for category, metrics_list in METRICAS_INE.items():
        for metric_tuple in metrics_list:
            metric_name = metric_tuple[0]
            metric_desc = metric_tuple[1] if isinstance(metric_tuple, tuple) else ""
            
            # Buscar correspondencias
            correspondences = find_metric_correspondence(metric_name, extracted_data)
            
            if correspondences:
                # Agrupar por nombre de métrica extraída
                metric_groups = {}
                for table_code, extracted_name in correspondences:
                    if extracted_name not in metric_groups:
                        metric_groups[extracted_name] = []
                    metric_groups[extracted_name].append(table_code)
                
                for extracted_name, table_codes in metric_groups.items():
                    rows.append({
                        'Categoría INE': category.replace('_', ' ').title(),
                        'Métrica INE': metric_name,
                        'Descripción INE': metric_desc[:80] + '...' if len(metric_desc) > 80 else metric_desc,
                        'Métrica Extraída': extracted_name,
                        'Tablas Origen': ', '.join(sorted(table_codes)),
                        'Nº Tablas': len(table_codes),
                        'Estado': '✓ Encontrada'
                    })
            else:
                rows.append({
                    'Categoría INE': category.replace('_', ' ').title(),
                    'Métrica INE': metric_name,
                    'Descripción INE': metric_desc[:80] + '...' if len(metric_desc) > 80 else metric_desc,
                    'Métrica Extraída': '-',
                    'Tablas Origen': '-',
                    'Nº Tablas': 0,
                    'Estado': '✗ No encontrada'
                })
    
    return pd.DataFrame(rows)

def save_to_excel(df: pd.DataFrame, output_file: Path):
    """Guarda el DataFrame en Excel con formato."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Correspondencia Métricas"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    found_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    not_found_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir encabezados
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir datos
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Colorear según estado
            if col_idx == 7:  # Columna Estado
                if '✓' in str(value):
                    cell.fill = found_fill
                elif '✗' in str(value):
                    cell.fill = not_found_fill
            
            # Alineación
            if col_idx in [6, 7]:  # Nº Tablas y Estado
                cell.alignment = Alignment(horizontal='center')
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20  # Categoría
    ws.column_dimensions['B'].width = 40  # Métrica INE
    ws.column_dimensions['C'].width = 50  # Descripción
    ws.column_dimensions['D'].width = 40  # Métrica Extraída
    ws.column_dimensions['E'].width = 30  # Tablas Origen
    ws.column_dimensions['F'].width = 12  # Nº Tablas
    ws.column_dimensions['G'].width = 15  # Estado
    
    # Agregar hoja de resumen
    ws_resumen = wb.create_sheet("Resumen")
    
    # Resumen por categoría
    resumen_data = df.groupby('Categoría INE').agg({
        'Estado': lambda x: sum('✓' in str(v) for v in x),
        'Métrica INE': 'count'
    }).rename(columns={'Estado': 'Encontradas', 'Métrica INE': 'Total INE'})
    resumen_data['Porcentaje'] = (resumen_data['Encontradas'] / resumen_data['Total INE'] * 100).round(1)
    
    # Escribir resumen
    ws_resumen.cell(row=1, column=1, value="RESUMEN DE COBERTURA")
    ws_resumen.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    headers = ['Categoría', 'Total INE', 'Encontradas', 'Porcentaje (%)']
    for col, header in enumerate(headers, 1):
        cell = ws_resumen.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 4
    for category, data in resumen_data.iterrows():
        ws_resumen.cell(row=row, column=1, value=category).border = border
        ws_resumen.cell(row=row, column=2, value=int(data['Total INE'])).border = border
        ws_resumen.cell(row=row, column=3, value=int(data['Encontradas'])).border = border
        cell = ws_resumen.cell(row=row, column=4, value=f"{data['Porcentaje']:.1f}%")
        cell.border = border
        if data['Porcentaje'] >= 80:
            cell.fill = found_fill
        elif data['Porcentaje'] < 50:
            cell.fill = not_found_fill
        row += 1
    
    # Totales
    total_ine = resumen_data['Total INE'].sum()
    total_found = resumen_data['Encontradas'].sum()
    total_pct = (total_found / total_ine * 100) if total_ine > 0 else 0
    
    ws_resumen.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_resumen.cell(row=row, column=2, value=int(total_ine)).font = Font(bold=True)
    ws_resumen.cell(row=row, column=3, value=int(total_found)).font = Font(bold=True)
    ws_resumen.cell(row=row, column=4, value=f"{total_pct:.1f}%").font = Font(bold=True)
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 15
    
    wb.save(output_file)

def main():
    """Función principal."""
    print("=" * 80)
    print("GENERANDO TABLA DE CORRESPONDENCIA MÉTRICAS INE vs EXTRAÍDAS")
    print("=" * 80)
    
    # Cargar datos extraídos
    print("\n1. Cargando métricas extraídas...")
    extracted_data = load_extracted_metrics()
    
    # Contar métricas únicas extraídas
    all_extracted_metrics = set()
    for table_data in extracted_data.values():
        if table_data and 'metricas' in table_data:
            for metric in table_data['metricas']:
                all_extracted_metrics.add(metric['nombre'])
    
    print(f"   - Total tablas procesadas: {len(extracted_data)}")
    print(f"   - Total métricas únicas extraídas: {len(all_extracted_metrics)}")
    
    # Generar tabla de correspondencia
    print("\n2. Generando tabla de correspondencia...")
    df_correspondence = generate_correspondence_table(extracted_data)
    
    # Guardar resultados
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Excel
    excel_file = REPORTS_DIR / f"correspondencia_metricas_INE_{timestamp}.xlsx"
    save_to_excel(df_correspondence, excel_file)
    print(f"\n3. Tabla Excel guardada en: {excel_file}")
    
    # CSV
    csv_file = REPORTS_DIR / f"correspondencia_metricas_INE_{timestamp}.csv"
    df_correspondence.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"   Tabla CSV guardada en: {csv_file}")
    
    # Mostrar resumen
    print("\n" + "=" * 80)
    print("RESUMEN DE CORRESPONDENCIAS")
    print("=" * 80)
    
    total_ine_metrics = len(df_correspondence)
    found_metrics = len(df_correspondence[df_correspondence['Estado'].str.contains('✓')])
    not_found_metrics = total_ine_metrics - found_metrics
    coverage_pct = (found_metrics / total_ine_metrics * 100) if total_ine_metrics > 0 else 0
    
    print(f"\nMétricas oficiales INE: {total_ine_metrics}")
    print(f"Métricas encontradas:    {found_metrics} ({coverage_pct:.1f}%)")
    print(f"Métricas no encontradas: {not_found_metrics} ({100-coverage_pct:.1f}%)")
    
    # Mostrar métricas no encontradas
    if not_found_metrics > 0:
        print("\nMétricas INE no encontradas:")
        not_found = df_correspondence[df_correspondence['Estado'].str.contains('✗')]
        for _, row in not_found.iterrows():
            print(f"  - [{row['Categoría INE']}] {row['Métrica INE']}")
    
    print("\n✅ Proceso completado exitosamente")

if __name__ == "__main__":
    main()