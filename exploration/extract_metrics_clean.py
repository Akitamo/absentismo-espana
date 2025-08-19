#!/usr/bin/env python3
"""
Script mejorado para extraer métricas limpias y depuradas de las 35 tablas del INE.
Separa correctamente métricas, unidades de medida, tipos de datos y dimensiones.
"""

import pandas as pd
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set, Tuple, Optional
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de rutas
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data" / "raw" / "csv"
CONFIG_FILE = BASE_DIR / "config" / "tables.json"
REPORTS_DIR = BASE_DIR / "data" / "exploration_reports"

# LISTAS DE EXCLUSIÓN - Valores EXACTOS que NO son métricas
EXCLUDE_EXACT_VALUES = {
    # Valores que deben coincidir exactamente (no como substring)
    'euros', 'euro', 'índice', 'indice', 'porcentaje', '%', 'horas', 'número',
    'total', 'totales', 'tipo de dato', 'principales series de etcl',
    'clase de indicador', 'componentes del coste', 'tiempo de trabajo',
    'ambas jornadas', 'jornada completa', 'jornada parcial', 
    'tiempo completo', 'tiempo parcial', 'tipo de jornada',
    'nan', 'none', 'null', '', ' ',
    'tasa de variación', 'tasa de variacion', 'tasa de variación anual',
    'base 100', 'base 2000', 'base 2008'
}

# Nombres de columnas que contienen las métricas reales
METRIC_COLUMN_NAMES = {
    'componentes del coste', 'componente del coste', 'componentes',
    'tiempo de trabajo', 'tiempos de trabajo',
    'motivos', 'motivo', 'causa', 'motivos de no vacantes',
    'variable', 'indicador', 'concepto', 'tipos'
}

# Nombres de columnas que contienen unidades/tipos
UNIT_COLUMN_NAMES = {
    'tipo de dato', 'unidad', 'unidad de medida', 'medida'
}

# Dimensiones conocidas (no métricas)
DIMENSION_COLUMN_NAMES = {
    'periodo', 'sectores', 'sector', 'actividad', 'secciones', 'divisiones',
    'comunidades', 'ccaa', 'comunidad autónoma', 'autonoma',
    'tipo de jornada', 'jornada', 'tamaño', 'establecimiento',
    'cnae', 'clase de indicador'
}

# Categorías de métricas según metodología INE
METRIC_CATEGORIES = {
    'COSTES_LABORALES': {
        'keywords': ['coste', 'cotización', 'cotizacion', 'subvención', 'subvencion',
                    'despido', 'prestación', 'prestacion', 'percepción', 'percepcion',
                    'bonificación', 'bonificacion', 'contingencia', 'fogasa'],
        'exclude': ['coste salarial']  # Estos van a COSTE_SALARIAL
    },
    'COSTE_SALARIAL': {
        'keywords': ['salarial', 'salario', 'ordinario', 'extraordinario', 'atrasado'],
        'exclude': []
    },
    'TIEMPO_TRABAJO': {
        'keywords': ['hora', 'tiempo', 'jornada', 'efectiva', 'pactada', 'pagada', 
                    'extra', 'extraordinaria', 'no trabajada', 'i.t.', 'it', 
                    'incapacidad', 'vacaciones', 'fiestas'],
        'exclude': ['hora extraordinaria']  # Si es coste por hora extra
    },
    'VACANTES': {
        'keywords': ['vacante', 'puesto', 'plaza'],
        'exclude': ['motivo', 'no vacante']
    },
    'MOTIVOS_NO_VACANTES': {
        'keywords': ['motivo', 'razón', 'razon', 'causa', 'no vacante', 'no hay vacante'],
        'exclude': []
    },
    'INDICES_SERIES': {
        'keywords': ['serie', 'variación', 'variacion', 'evolución', 'evolucion'],
        'exclude': []
    }
}

def load_config() -> Dict:
    """Carga la configuración de tablas desde el archivo JSON."""
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Aplanar estructura de categorías
    tables = {}
    for categoria_key, categoria_data in data.get('categorias', {}).items():
        for table_code, table_info in categoria_data.get('tablas', {}).items():
            table_info['categoria'] = categoria_data.get('descripcion', categoria_key)
            tables[table_code] = table_info
    
    return tables

def detect_encoding(file_path: Path) -> str:
    """Detecta la codificación del archivo CSV."""
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                f.read(1000)
            return encoding
        except UnicodeDecodeError:
            continue
    return 'utf-8'

def clean_value(value: str) -> str:
    """Limpia y normaliza un valor."""
    if pd.isna(value):
        return ''
    return str(value).strip().lower()

def is_excluded_value(value: str) -> bool:
    """Verifica si un valor debe ser excluido (no es métrica)."""
    value_clean = clean_value(value)
    
    # Verificar coincidencia exacta con valores excluidos
    if value_clean in EXCLUDE_EXACT_VALUES:
        return True
    
    # Verificar si es solo números o caracteres especiales
    if value_clean.replace('.', '').replace(',', '').replace('-', '').isdigit():
        return True
    
    # Verificar si tiene menos de 3 caracteres (probablemente no es métrica válida)
    if len(value_clean) < 3:
        return True
    
    return False

def categorize_metric(metric_name: str) -> str:
    """Categoriza una métrica según su nombre y keywords."""
    metric_lower = metric_name.lower()
    
    for category, config in METRIC_CATEGORIES.items():
        # Verificar exclusiones primero
        excluded = False
        for exclude_term in config.get('exclude', []):
            if exclude_term in metric_lower:
                excluded = True
                break
        
        if not excluded:
            # Verificar keywords
            for keyword in config.get('keywords', []):
                if keyword in metric_lower:
                    return category
    
    return 'OTROS'

def extract_table_metrics_clean(table_code: str, table_info: Dict) -> Dict:
    """Extrae métricas limpias de una tabla específica."""
    # Buscar archivo CSV
    csv_pattern = f"{table_code}_*.csv"
    csv_files = list(DATA_DIR.glob(csv_pattern))
    
    if not csv_files:
        print(f"  [!] No se encontró archivo para tabla {table_code}")
        return None
    
    csv_file = csv_files[0]
    print(f"  Analizando: {csv_file.name}")
    
    # Detectar encoding y leer archivo
    encoding = detect_encoding(csv_file)
    
    try:
        # Leer CSV
        df = pd.read_csv(csv_file, encoding=encoding, sep=';', decimal=',')
        
        if df is None or len(df) == 0:
            print(f"  [!] No se pudo leer el archivo {csv_file.name}")
            return None
        
        # Identificar columnas
        columns_lower = {col: col.lower().strip() for col in df.columns}
        
        # Buscar columna de métricas
        metric_column = None
        unit_column = None
        dimensions = []
        
        for col, col_lower in columns_lower.items():
            # Identificar columna de métricas
            if any(metric_name in col_lower for metric_name in METRIC_COLUMN_NAMES):
                metric_column = col
            # Identificar columna de unidades
            elif any(unit_name in col_lower for unit_name in UNIT_COLUMN_NAMES):
                unit_column = col
            # Identificar dimensiones
            elif any(dim_name in col_lower for dim_name in DIMENSION_COLUMN_NAMES):
                dimensions.append(col)
            # La columna 'Total' contiene valores, no es dimensión ni métrica
            elif col_lower in ['total', 'valor']:
                pass  # Columna de valores numéricos
        
        # Extraer métricas únicas
        metrics = []
        metrics_set = set()  # Para evitar duplicados
        
        if metric_column:
            # Formato largo: métricas en una columna
            unique_metrics = df[metric_column].dropna().unique()
            
            for metric_name in unique_metrics:
                metric_clean = str(metric_name).strip()
                
                # Filtrar valores excluidos
                if not is_excluded_value(metric_clean) and metric_clean not in metrics_set:
                    metrics_set.add(metric_clean)
                    
                    # Determinar unidad de medida
                    unit = "No especificada"
                    if unit_column and unit_column in df.columns:
                        # Buscar la unidad asociada a esta métrica
                        metric_rows = df[df[metric_column] == metric_name]
                        if not metric_rows.empty and unit_column in metric_rows.columns:
                            units = metric_rows[unit_column].dropna().unique()
                            if len(units) > 0:
                                unit = str(units[0])
                    
                    # Inferir unidad del nombre de la tabla si no se encontró
                    if unit == "No especificada":
                        if "por hora" in table_info.get('nombre', '').lower():
                            unit = "Euros/hora"
                        elif "por trabajador" in table_info.get('nombre', '').lower():
                            unit = "Euros/trabajador-mes"
                        elif "tiempo" in table_info.get('nombre', '').lower():
                            unit = "Horas"
                        elif "vacante" in table_info.get('nombre', '').lower():
                            unit = "Número"
                        elif "motivo" in table_info.get('nombre', '').lower():
                            unit = "Porcentaje"
                    
                    metrics.append({
                        'nombre': metric_clean,
                        'categoria': categorize_metric(metric_clean),
                        'unidad_medida': unit,
                        'columna_origen': metric_column
                    })
        
        # Si no hay columna de métricas, buscar en los nombres de columnas (formato ancho)
        if not metrics and not metric_column:
            for col in df.columns:
                col_lower = col.lower().strip()
                # Saltar dimensiones conocidas y columna de valores
                if col_lower not in ['total', 'periodo'] and not any(dim in col_lower for dim in DIMENSION_COLUMN_NAMES):
                    # Verificar si es columna numérica
                    try:
                        pd.to_numeric(df[col].dropna().iloc[:10], errors='raise')
                        # Es una métrica en formato columna
                        if not is_excluded_value(col) and col not in metrics_set:
                            metrics_set.add(col)
                            metrics.append({
                                'nombre': col,
                                'categoria': categorize_metric(col),
                                'unidad_medida': "Inferir del contexto",
                                'columna_origen': 'columna_propia'
                            })
                    except:
                        pass
        
        # Crear resultado estructurado
        result = {
            'codigo': table_code,
            'nombre': table_info.get('nombre', ''),
            'categoria_tabla': table_info.get('categoria', ''),
            'archivo': csv_file.name,
            'formato_datos': 'largo' if metric_column else 'ancho',
            'columna_metricas': metric_column if metric_column else 'N/A',
            'columna_unidades': unit_column if unit_column else 'N/A',
            'dimensiones': dimensions,
            'total_dimensiones': len(dimensions),
            'metricas': metrics,
            'total_metricas': len(metrics),
            'metricas_por_categoria': {}
        }
        
        # Agrupar métricas por categoría
        for categoria in METRIC_CATEGORIES.keys():
            metricas_categoria = [m for m in metrics if m['categoria'] == categoria]
            if metricas_categoria:
                result['metricas_por_categoria'][categoria] = {
                    'total': len(metricas_categoria),
                    'metricas': [m['nombre'] for m in metricas_categoria]
                }
        
        # Agregar categoría "OTROS" si hay métricas sin categorizar
        otros = [m for m in metrics if m['categoria'] == 'OTROS']
        if otros:
            result['metricas_por_categoria']['OTROS'] = {
                'total': len(otros),
                'metricas': [m['nombre'] for m in otros]
            }
        
        return result
        
    except Exception as e:
        print(f"  [!] Error procesando {csv_file.name}: {str(e)}")
        return None

def generate_excel_report_clean(all_metrics: Dict, output_file: Path):
    """Genera un informe Excel limpio y estructurado."""
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # HOJA 1: Catálogo de Métricas Únicas
    ws_catalogo = wb.active
    ws_catalogo.title = "Catálogo Métricas"
    
    # Recopilar todas las métricas únicas
    all_unique_metrics = {}
    for table_code, data in all_metrics.items():
        if data and data['metricas']:
            for metric in data['metricas']:
                metric_name = metric['nombre']
                if metric_name not in all_unique_metrics:
                    all_unique_metrics[metric_name] = {
                        'nombre': metric_name,
                        'categoria': metric['categoria'],
                        'unidades': set(),
                        'tablas': []
                    }
                all_unique_metrics[metric_name]['unidades'].add(metric['unidad_medida'])
                all_unique_metrics[metric_name]['tablas'].append(table_code)
    
    # Encabezados catálogo
    headers_catalogo = ["Métrica", "Categoría", "Unidades de Medida", "Nº Tablas", "Tablas donde aparece"]
    for col, header in enumerate(headers_catalogo, 1):
        cell = ws_catalogo.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos catálogo
    row = 2
    for metric_name in sorted(all_unique_metrics.keys()):
        metric_info = all_unique_metrics[metric_name]
        ws_catalogo.cell(row=row, column=1, value=metric_name).border = border
        ws_catalogo.cell(row=row, column=2, value=metric_info['categoria']).border = border
        ws_catalogo.cell(row=row, column=3, value=', '.join(metric_info['unidades'])).border = border
        ws_catalogo.cell(row=row, column=4, value=len(metric_info['tablas'])).border = border
        ws_catalogo.cell(row=row, column=5, value=', '.join(sorted(metric_info['tablas']))).border = border
        row += 1
    
    # Ajustar anchos
    ws_catalogo.column_dimensions['A'].width = 60
    ws_catalogo.column_dimensions['B'].width = 25
    ws_catalogo.column_dimensions['C'].width = 30
    ws_catalogo.column_dimensions['D'].width = 12
    ws_catalogo.column_dimensions['E'].width = 40
    
    # HOJA 2: Resumen por Tabla
    ws_resumen = wb.create_sheet("Resumen por Tabla")
    
    headers_resumen = ["Código", "Nombre Tabla", "Formato", "Total Métricas", 
                      "Costes Lab.", "Coste Sal.", "Tiempo Trab.", "Vacantes", 
                      "Motivos NV", "Series", "Otros"]
    
    for col, header in enumerate(headers_resumen, 1):
        cell = ws_resumen.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 2
    for table_code, data in sorted(all_metrics.items()):
        if data:
            ws_resumen.cell(row=row, column=1, value=table_code).border = border
            ws_resumen.cell(row=row, column=2, value=data['nombre'][:50]).border = border
            ws_resumen.cell(row=row, column=3, value=data['formato_datos']).border = border
            ws_resumen.cell(row=row, column=4, value=data['total_metricas']).border = border
            
            col = 5
            for cat in ['COSTES_LABORALES', 'COSTE_SALARIAL', 'TIEMPO_TRABAJO', 
                       'VACANTES', 'MOTIVOS_NO_VACANTES', 'INDICES_SERIES', 'OTROS']:
                valor = data['metricas_por_categoria'].get(cat, {}).get('total', 0)
                ws_resumen.cell(row=row, column=col, value=valor).border = border
                col += 1
            
            row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 10
    ws_resumen.column_dimensions['B'].width = 50
    ws_resumen.column_dimensions['C'].width = 12
    for col in range(4, 12):
        ws_resumen.column_dimensions[get_column_letter(col)].width = 12
    
    # HOJA 3: Detalle por Tabla
    ws_detalle = wb.create_sheet("Detalle por Tabla")
    
    headers_detalle = ["Código Tabla", "Nombre Tabla", "Métrica", "Categoría", "Unidad Medida"]
    
    for col, header in enumerate(headers_detalle, 1):
        cell = ws_detalle.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 2
    for table_code, data in sorted(all_metrics.items()):
        if data and data['metricas']:
            # Agrupar por categoría
            by_category = {}
            for metrica in data['metricas']:
                cat = metrica['categoria']
                if cat not in by_category:
                    by_category[cat] = []
                by_category[cat].append(metrica)
            
            # Escribir por categoría
            for categoria in sorted(by_category.keys()):
                for metrica in by_category[categoria]:
                    ws_detalle.cell(row=row, column=1, value=table_code).border = border
                    ws_detalle.cell(row=row, column=2, value=data['nombre'][:50]).border = border
                    ws_detalle.cell(row=row, column=3, value=metrica['nombre']).border = border
                    ws_detalle.cell(row=row, column=4, value=metrica['categoria']).border = border
                    ws_detalle.cell(row=row, column=5, value=metrica['unidad_medida']).border = border
                    row += 1
    
    # Ajustar anchos
    ws_detalle.column_dimensions['A'].width = 12
    ws_detalle.column_dimensions['B'].width = 50
    ws_detalle.column_dimensions['C'].width = 60
    ws_detalle.column_dimensions['D'].width = 20
    ws_detalle.column_dimensions['E'].width = 25
    
    # HOJA 4: Matriz de Presencia
    ws_matriz = wb.create_sheet("Matriz Presencia")
    
    # Preparar matriz
    table_codes = sorted([code for code in all_metrics.keys() if all_metrics[code]])
    metric_names = sorted(list(all_unique_metrics.keys()))
    
    # Encabezado con códigos de tabla
    ws_matriz.cell(row=1, column=1, value="Métrica \\ Tabla")
    ws_matriz.cell(row=1, column=1).fill = header_fill
    ws_matriz.cell(row=1, column=1).font = header_font
    
    for col, table_code in enumerate(table_codes, 2):
        cell = ws_matriz.cell(row=1, column=col, value=table_code)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(text_rotation=90, horizontal='center')
    
    # Llenar matriz
    for row, metric_name in enumerate(metric_names, 2):
        ws_matriz.cell(row=row, column=1, value=metric_name[:50])
        ws_matriz.cell(row=row, column=1).border = border
        
        for col, table_code in enumerate(table_codes, 2):
            if table_code in all_unique_metrics[metric_name]['tablas']:
                cell = ws_matriz.cell(row=row, column=col, value="X")
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                ws_matriz.cell(row=row, column=col, value="")
            ws_matriz.cell(row=row, column=col).border = border
    
    # Ajustar anchos
    ws_matriz.column_dimensions['A'].width = 50
    for col in range(2, len(table_codes) + 2):
        ws_matriz.column_dimensions[get_column_letter(col)].width = 4
    
    # Guardar archivo
    wb.save(output_file)
    print(f"\n[OK] Informe Excel limpio guardado en: {output_file}")

def main():
    """Función principal."""
    print("=" * 80)
    print("EXTRACCIÓN LIMPIA Y DEPURADA DE MÉTRICAS INE")
    print("=" * 80)
    
    # Cargar configuración
    config = load_config()
    
    # Procesar todas las tablas
    all_metrics = {}
    total_processed = 0
    total_metrics_found = 0
    
    for table_code, table_info in config.items():
        print(f"\nProcesando tabla {table_code}: {table_info['nombre'][:50]}...")
        
        result = extract_table_metrics_clean(table_code, table_info)
        
        if result:
            all_metrics[table_code] = result
            total_processed += 1
            total_metrics_found += result['total_metricas']
            print(f"  -> {result['total_metricas']} métricas limpias encontradas")
            print(f"  -> Formato: {result['formato_datos']}")
    
    # Generar timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Guardar JSON detallado
    json_file = REPORTS_DIR / f"metricas_limpias_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(all_metrics, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n[OK] Datos JSON guardados en: {json_file}")
    
    # Generar Excel
    excel_file = REPORTS_DIR / f"metricas_limpias_{timestamp}.xlsx"
    generate_excel_report_clean(all_metrics, excel_file)
    
    # Resumen final
    print("\n" + "=" * 80)
    print("RESUMEN FINAL")
    print("=" * 80)
    print(f"Tablas procesadas: {total_processed}/{len(config)}")
    print(f"Total métricas limpias encontradas: {total_metrics_found}")
    
    # Contar métricas únicas
    unique_metrics = set()
    category_totals = {}
    
    for data in all_metrics.values():
        if data:
            for metric in data['metricas']:
                unique_metrics.add(metric['nombre'])
                cat = metric['categoria']
                category_totals[cat] = category_totals.get(cat, 0) + 1
    
    print(f"Métricas únicas identificadas: {len(unique_metrics)}")
    
    print("\nMétricas por categoría:")
    for cat in sorted(category_totals.keys()):
        print(f"  - {cat}: {category_totals[cat]}")
    
    print("\n[OK] Proceso completado exitosamente")
    print(f"[OK] Revise el archivo Excel para el catálogo completo de métricas")

if __name__ == "__main__":
    main()