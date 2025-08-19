#!/usr/bin/env python3
"""
Script MEJORADO para extraer métricas de las 35 tablas del INE.
Soluciona problemas detectados en la validación:
1. Detecta tablas simples de 3 columnas (vacantes)
2. Maneja typos en nombres de columnas
3. Evita exclusiones incorrectas
4. Mejora normalización para matching
5. Identifica métricas calculadas
"""

import pandas as pd
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set, Tuple, Optional
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de rutas
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data" / "raw" / "csv"
CONFIG_FILE = BASE_DIR / "config" / "tables.json"
REPORTS_DIR = BASE_DIR / "data" / "exploration_reports"

# LISTAS DE EXCLUSIÓN MEJORADAS - Más específicas
EXCLUDE_EXACT_VALUES = {
    # Solo excluir valores exactos que definitivamente no son métricas
    'total', 'totales', 'tipo de dato', 'principales series de etcl',
    'clase de indicador', 
    'nan', 'none', 'null', '', ' '
}

# Valores que NO excluir aunque sean cortos si vienen de columna de métricas
NEVER_EXCLUDE_FROM_METRIC_COLUMN = {
    'euros', 'horas', 'número', 'índice', 'porcentaje'
}

# Nombres de columnas que contienen métricas (con typos comunes)
METRIC_COLUMN_NAMES = {
    'componentes del coste', 'componente del coste', 'componentes',
    'tiempo de trabajo', 'tiempos de trabajo',
    'motivos', 'motivo', 'causa', 'razón', 'razon',
    'motivos por los que no exiten', 'motivos por los que no existen',  # Typo común
    'variable', 'indicador', 'concepto', 'tipos', 'tipo'
}

# Columnas que son dimensiones, no métricas
DIMENSION_COLUMN_NAMES = {
    'periodo', 'sectores', 'sector', 'actividad', 'secciones', 'divisiones',
    'comunidades', 'ccaa', 'comunidad autónoma', 'autonoma',
    'tipo de jornada', 'jornada', 'tamaño', 'establecimiento',
    'cnae', 'clase de indicador'
}

# Métricas que el INE calcula pero no aparecen directamente en CSVs
CALCULATED_METRICS = {
    'Percepciones por día de I.T.',
    'Coste indemnización trabajador despedido',
    'Coste por hora extra'
}

# Métricas implícitas en tablas de 3 columnas
IMPLICIT_METRICS = {
    '6047': 'Número de vacantes',
    '6048': 'Número de vacantes',
    '6049': 'Número de vacantes',
    '6064': 'Número de vacantes'
}

# Categorías de métricas según metodología INE
METRIC_CATEGORIES = {
    'COSTES_LABORALES': {
        'keywords': ['coste', 'cotización', 'cotizacion', 'subvención', 'subvencion',
                    'despido', 'prestación', 'prestacion', 'percepción', 'percepcion',
                    'bonificación', 'bonificacion', 'contingencia', 'fogasa'],
        'exclude': ['coste salarial']
    },
    'COSTE_SALARIAL': {
        'keywords': ['salarial', 'salario', 'ordinario', 'extraordinario', 'atrasado'],
        'exclude': []
    },
    'TIEMPO_TRABAJO': {
        'keywords': ['hora', 'tiempo', 'jornada', 'efectiva', 'pactada', 'pagada', 
                    'extra', 'extraordinaria', 'no trabajada', 'i.t.', 'it', 
                    'incapacidad', 'vacaciones', 'fiestas', 'maternidad'],
        'exclude': ['coste por hora']
    },
    'VACANTES': {
        'keywords': ['vacante', 'puesto', 'plaza'],
        'exclude': ['motivo', 'no vacante']
    },
    'MOTIVOS_NO_VACANTES': {
        'keywords': ['motivo', 'razón', 'razon', 'causa', 'no vacante', 
                     'no exiten', 'no existen', 'elevado coste', 'no se necesita'],
        'exclude': []
    },
    'INDICES_SERIES': {
        'keywords': ['serie', 'variación', 'variacion', 'evolución', 'evolucion'],
        'exclude': []
    }
}

def load_config() -> Dict:
    """Carga la configuración de tablas desde el archivo JSON."""
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Aplanar estructura de categorías
    tables = {}
    for categoria_key, categoria_data in data.get('categorias', {}).items():
        for table_code, table_info in categoria_data.get('tablas', {}).items():
            table_info['categoria'] = categoria_data.get('descripcion', categoria_key)
            tables[table_code] = table_info
    
    return tables

def detect_encoding(file_path: Path) -> str:
    """Detecta la codificación del archivo CSV."""
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                f.read(1000)
            return encoding
        except UnicodeDecodeError:
            continue
    return 'utf-8'

def normalize_for_comparison(text: str) -> str:
    """Normaliza texto para comparación flexible."""
    if pd.isna(text):
        return ''
    normalized = str(text).lower().strip()
    # Normalizar I.T. vs I.T vs IT
    normalized = normalized.replace('i.t.', 'it')
    normalized = normalized.replace('i.t', 'it')
    # Normalizar S.Social
    normalized = normalized.replace('s.social', 'seguridad social')
    # Normalizar espacios y caracteres
    normalized = normalized.replace('.', '').replace(',', '')
    normalized = ' '.join(normalized.split())
    return normalized

def is_excluded_value(value: str, from_metric_column: bool = False) -> bool:
    """Verifica si un valor debe ser excluido (no es métrica)."""
    value_clean = normalize_for_comparison(value)
    
    # Si viene de una columna de métricas identificada, ser menos estricto
    if from_metric_column:
        # No excluir valores específicos aunque sean cortos
        if value_clean in NEVER_EXCLUDE_FROM_METRIC_COLUMN:
            return False
        # Solo excluir si está en la lista exacta de exclusión
        if value_clean in EXCLUDE_EXACT_VALUES:
            return True
        # No excluir por longitud si viene de columna de métricas
        return False
    
    # Para otras columnas, aplicar reglas normales
    if value_clean in EXCLUDE_EXACT_VALUES:
        return True
    
    # Verificar si es solo números
    if value_clean.replace('.', '').replace(',', '').replace('-', '').isdigit():
        return True
    
    # Verificar longitud mínima solo si NO es de columna de métricas
    if len(value_clean) < 3:
        return True
    
    return False

def categorize_metric(metric_name: str) -> str:
    """Categoriza una métrica según su nombre y keywords."""
    metric_lower = normalize_for_comparison(metric_name)
    
    for category, config in METRIC_CATEGORIES.items():
        # Verificar exclusiones primero
        excluded = False
        for exclude_term in config.get('exclude', []):
            if exclude_term in metric_lower:
                excluded = True
                break
        
        if not excluded:
            # Verificar keywords
            for keyword in config.get('keywords', []):
                if keyword in metric_lower:
                    return category
    
    return 'OTROS'

def extract_implicit_metrics(table_code: str, df: pd.DataFrame) -> List[Dict]:
    """Extrae métricas implícitas de tablas simples (3 columnas)."""
    metrics = []
    
    # Verificar si es una tabla de 3 columnas con métrica implícita
    if len(df.columns) == 3 and table_code in IMPLICIT_METRICS:
        metric_name = IMPLICIT_METRICS[table_code]
        metrics.append({
            'nombre': metric_name,
            'categoria': categorize_metric(metric_name),
            'unidad_medida': 'Número' if 'vacante' in metric_name.lower() else 'No especificada',
            'columna_origen': 'implícita',
            'tipo': 'directa'
        })
        print(f"    -> Métrica implícita detectada: {metric_name}")
    
    return metrics

def extract_table_metrics_enhanced(table_code: str, table_info: Dict) -> Dict:
    """Extrae métricas con mejoras para casos especiales."""
    # Buscar archivo CSV
    csv_pattern = f"{table_code}_*.csv"
    csv_files = list(DATA_DIR.glob(csv_pattern))
    
    if not csv_files:
        print(f"  [!] No se encontró archivo para tabla {table_code}")
        return None
    
    csv_file = csv_files[0]
    print(f"  Analizando: {csv_file.name}")
    
    # Detectar encoding y leer archivo
    encoding = detect_encoding(csv_file)
    
    try:
        # Leer CSV
        df = pd.read_csv(csv_file, encoding=encoding, sep=';', decimal=',')
        
        if df is None or len(df) == 0:
            print(f"  [!] No se pudo leer el archivo {csv_file.name}")
            return None
        
        # Identificar columnas
        columns_lower = {col: normalize_for_comparison(col) for col in df.columns}
        
        # Buscar columna de métricas (con tolerancia a typos)
        metric_column = None
        unit_column = None
        dimensions = []
        
        for col, col_lower in columns_lower.items():
            # Buscar columna de métricas con matching flexible
            is_metric_col = False
            for metric_name in METRIC_COLUMN_NAMES:
                # Matching parcial para manejar typos
                if metric_name in col_lower or col_lower in metric_name:
                    metric_column = col
                    is_metric_col = True
                    print(f"    -> Columna de métricas encontrada: {col}")
                    break
            
            if not is_metric_col:
                # Identificar columna de unidades
                if any(unit_name in col_lower for unit_name in ['tipo de dato', 'unidad', 'medida']):
                    unit_column = col
                # Identificar dimensiones
                elif any(dim_name in col_lower for dim_name in DIMENSION_COLUMN_NAMES):
                    dimensions.append(col)
        
        # Extraer métricas
        metrics = []
        metrics_set = set()
        
        # 1. Extraer métricas explícitas de columnas identificadas
        if metric_column:
            unique_metrics = df[metric_column].dropna().unique()
            print(f"    -> Valores únicos encontrados: {len(unique_metrics)}")
            
            for metric_name in unique_metrics:
                metric_clean = str(metric_name).strip()
                metric_normalized = normalize_for_comparison(metric_clean)
                
                # Usar exclusión menos estricta para columnas de métricas
                if not is_excluded_value(metric_clean, from_metric_column=True) and metric_normalized not in metrics_set:
                    metrics_set.add(metric_normalized)
                    
                    # Determinar unidad
                    unit = "No especificada"
                    if unit_column and unit_column in df.columns:
                        metric_rows = df[df[metric_column] == metric_name]
                        if not metric_rows.empty and unit_column in metric_rows.columns:
                            units = metric_rows[unit_column].dropna().unique()
                            if len(units) > 0:
                                unit = str(units[0])
                    
                    # Inferir unidad del contexto si no se encontró
                    if unit == "No especificada":
                        if "hora" in table_info.get('nombre', '').lower():
                            unit = "Euros/hora" if "coste" in metric_clean.lower() else "Horas"
                        elif "trabajador" in table_info.get('nombre', '').lower():
                            unit = "Euros/trabajador-mes" if "coste" in metric_clean.lower() else "Por trabajador"
                        elif "vacante" in metric_clean.lower():
                            unit = "Número"
                        elif "motivo" in metric_clean.lower():
                            unit = "Porcentaje"
                    
                    metrics.append({
                        'nombre': metric_clean,
                        'categoria': categorize_metric(metric_clean),
                        'unidad_medida': unit,
                        'columna_origen': metric_column,
                        'tipo': 'directa'
                    })
        
        # 2. Detectar métricas implícitas en tablas simples
        implicit_metrics = extract_implicit_metrics(table_code, df)
        for im in implicit_metrics:
            if normalize_for_comparison(im['nombre']) not in metrics_set:
                metrics.append(im)
                metrics_set.add(normalize_for_comparison(im['nombre']))
        
        # 3. Agregar nota sobre métricas calculadas esperadas
        calculated_in_table = []
        table_name_lower = table_info.get('nombre', '').lower()
        if 'despido' in table_name_lower or 'indemnización' in table_name_lower:
            calculated_in_table.append('Coste indemnización trabajador despedido')
        if 'hora' in table_name_lower and 'extra' in table_name_lower:
            calculated_in_table.append('Coste por hora extra')
        if 'i.t' in table_name_lower or 'incapacidad' in table_name_lower:
            calculated_in_table.append('Percepciones por día de I.T.')
        
        # Crear resultado estructurado
        result = {
            'codigo': table_code,
            'nombre': table_info.get('nombre', ''),
            'categoria_tabla': table_info.get('categoria', ''),
            'archivo': csv_file.name,
            'formato_datos': 'largo' if metric_column else ('implícito' if implicit_metrics else 'ancho'),
            'columna_metricas': metric_column if metric_column else 'N/A',
            'columna_unidades': unit_column if unit_column else 'N/A',
            'dimensiones': dimensions,
            'total_dimensiones': len(dimensions),
            'metricas': metrics,
            'total_metricas': len(metrics),
            'metricas_calculadas_esperadas': calculated_in_table,
            'metricas_por_categoria': {}
        }
        
        # Agrupar métricas por categoría
        for categoria in METRIC_CATEGORIES.keys():
            metricas_categoria = [m for m in metrics if m['categoria'] == categoria]
            if metricas_categoria:
                result['metricas_por_categoria'][categoria] = {
                    'total': len(metricas_categoria),
                    'metricas': [m['nombre'] for m in metricas_categoria]
                }
        
        # Agregar categoría "OTROS"
        otros = [m for m in metrics if m['categoria'] == 'OTROS']
        if otros:
            result['metricas_por_categoria']['OTROS'] = {
                'total': len(otros),
                'metricas': [m['nombre'] for m in otros]
            }
        
        return result
        
    except Exception as e:
        print(f"  [!] Error procesando {csv_file.name}: {str(e)}")
        return None

def generate_excel_report_enhanced(all_metrics: Dict, output_file: Path):
    """Genera un informe Excel mejorado con información adicional."""
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # HOJA 1: Resumen mejorado
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Mejorado"
    
    # Título
    ws_resumen.merge_cells('A1:E1')
    ws_resumen['A1'] = "EXTRACCIÓN MEJORADA DE MÉTRICAS INE"
    ws_resumen['A1'].font = Font(size=14, bold=True)
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    # Estadísticas
    row = 3
    ws_resumen[f'A{row}'] = "ESTADÍSTICAS GENERALES"
    ws_resumen[f'A{row}'].font = Font(bold=True)
    row += 1
    
    total_tables = len(all_metrics)
    total_metrics = sum(data['total_metricas'] for data in all_metrics.values() if data)
    tables_with_implicit = sum(1 for data in all_metrics.values() if data and data['formato_datos'] == 'implícito')
    tables_with_calculated = sum(1 for data in all_metrics.values() if data and data.get('metricas_calculadas_esperadas'))
    
    ws_resumen[f'A{row}'] = "Total tablas procesadas:"
    ws_resumen[f'B{row}'] = total_tables
    row += 1
    ws_resumen[f'A{row}'] = "Total métricas encontradas:"
    ws_resumen[f'B{row}'] = total_metrics
    row += 1
    ws_resumen[f'A{row}'] = "Tablas con métricas implícitas:"
    ws_resumen[f'B{row}'] = tables_with_implicit
    ws_resumen[f'B{row}'].fill = success_fill if tables_with_implicit > 0 else warning_fill
    row += 1
    ws_resumen[f'A{row}'] = "Tablas con métricas calculadas esperadas:"
    ws_resumen[f'B{row}'] = tables_with_calculated
    row += 2
    
    # Resumen por categoría
    ws_resumen[f'A{row}'] = "MÉTRICAS POR CATEGORÍA"
    ws_resumen[f'A{row}'].font = Font(bold=True)
    row += 1
    
    category_totals = {}
    for data in all_metrics.values():
        if data:
            for cat, info in data['metricas_por_categoria'].items():
                category_totals[cat] = category_totals.get(cat, 0) + info['total']
    
    for cat, total in sorted(category_totals.items()):
        ws_resumen[f'A{row}'] = cat
        ws_resumen[f'B{row}'] = total
        row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 40
    ws_resumen.column_dimensions['B'].width = 20
    
    # HOJA 2: Detalle de métricas
    ws_detalle = wb.create_sheet("Detalle Métricas")
    
    headers = ["Código", "Tabla", "Métrica", "Categoría", "Unidad", "Tipo", "Origen"]
    for col, header in enumerate(headers, 1):
        cell = ws_detalle.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 2
    for table_code, data in sorted(all_metrics.items()):
        if data and data['metricas']:
            for metrica in data['metricas']:
                ws_detalle.cell(row=row, column=1, value=table_code).border = border
                ws_detalle.cell(row=row, column=2, value=data['nombre'][:40]).border = border
                ws_detalle.cell(row=row, column=3, value=metrica['nombre']).border = border
                ws_detalle.cell(row=row, column=4, value=metrica['categoria']).border = border
                ws_detalle.cell(row=row, column=5, value=metrica.get('unidad_medida', '')).border = border
                ws_detalle.cell(row=row, column=6, value=metrica.get('tipo', 'directa')).border = border
                ws_detalle.cell(row=row, column=7, value=metrica.get('columna_origen', '')).border = border
                
                # Colorear métricas implícitas
                if metrica.get('columna_origen') == 'implícita':
                    for col in range(1, 8):
                        ws_detalle.cell(row=row, column=col).fill = success_fill
                row += 1
    
    # Ajustar anchos
    ws_detalle.column_dimensions['A'].width = 10
    ws_detalle.column_dimensions['B'].width = 40
    ws_detalle.column_dimensions['C'].width = 50
    ws_detalle.column_dimensions['D'].width = 20
    ws_detalle.column_dimensions['E'].width = 25
    ws_detalle.column_dimensions['F'].width = 12
    ws_detalle.column_dimensions['G'].width = 20
    
    # HOJA 3: Métricas calculadas esperadas
    ws_calculadas = wb.create_sheet("Métricas Calculadas")
    
    ws_calculadas['A1'] = "MÉTRICAS CALCULADAS ESPERADAS (No en CSVs directamente)"
    ws_calculadas['A1'].font = Font(bold=True)
    
    row = 3
    headers = ["Tabla", "Métrica Calculada Esperada"]
    for col, header in enumerate(headers, 1):
        cell = ws_calculadas.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    for table_code, data in sorted(all_metrics.items()):
        if data and data.get('metricas_calculadas_esperadas'):
            for calc_metric in data['metricas_calculadas_esperadas']:
                ws_calculadas.cell(row=row, column=1, value=f"{table_code} - {data['nombre'][:40]}")
                ws_calculadas.cell(row=row, column=2, value=calc_metric)
                ws_calculadas.cell(row=row, column=2).fill = warning_fill
                row += 1
    
    ws_calculadas.column_dimensions['A'].width = 50
    ws_calculadas.column_dimensions['B'].width = 50
    
    # Guardar
    wb.save(output_file)
    print(f"\n[OK] Informe Excel mejorado guardado en: {output_file}")

def main():
    """Función principal."""
    print("=" * 80)
    print("EXTRACCIÓN MEJORADA DE MÉTRICAS INE")
    print("Versión 2.0 - Con detección de casos especiales")
    print("=" * 80)
    
    # Cargar configuración
    config = load_config()
    
    # Procesar todas las tablas
    all_metrics = {}
    total_processed = 0
    total_metrics_found = 0
    total_implicit = 0
    
    for table_code, table_info in config.items():
        print(f"\nProcesando tabla {table_code}: {table_info['nombre'][:50]}...")
        
        result = extract_table_metrics_enhanced(table_code, table_info)
        
        if result:
            all_metrics[table_code] = result
            total_processed += 1
            total_metrics_found += result['total_metricas']
            
            if result['formato_datos'] == 'implícito':
                total_implicit += 1
            
            print(f"  -> {result['total_metricas']} métricas encontradas")
            print(f"  -> Formato: {result['formato_datos']}")
            
            if result.get('metricas_calculadas_esperadas'):
                print(f"  -> Métricas calculadas esperadas: {len(result['metricas_calculadas_esperadas'])}")
    
    # Generar timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Guardar JSON
    json_file = REPORTS_DIR / f"metricas_enhanced_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(all_metrics, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n[OK] Datos JSON guardados en: {json_file}")
    
    # Generar Excel
    excel_file = REPORTS_DIR / f"metricas_enhanced_{timestamp}.xlsx"
    generate_excel_report_enhanced(all_metrics, excel_file)
    
    # Resumen final
    print("\n" + "=" * 80)
    print("RESUMEN FINAL - VERSIÓN MEJORADA")
    print("=" * 80)
    print(f"Tablas procesadas: {total_processed}/{len(config)}")
    print(f"Total métricas encontradas: {total_metrics_found}")
    print(f"Tablas con métricas implícitas detectadas: {total_implicit}")
    
    # Contar métricas únicas
    unique_metrics = set()
    for data in all_metrics.values():
        if data:
            for metric in data['metricas']:
                unique_metrics.add(normalize_for_comparison(metric['nombre']))
    
    print(f"Métricas únicas identificadas: {len(unique_metrics)}")
    
    # Resumen por categoría
    category_totals = {}
    for data in all_metrics.values():
        if data:
            for cat, info in data['metricas_por_categoria'].items():
                category_totals[cat] = category_totals.get(cat, 0) + info['total']
    
    print("\nMétricas por categoría:")
    for cat in sorted(category_totals.keys()):
        print(f"  - {cat}: {category_totals[cat]}")
    
    print("\n[OK] Proceso completado exitosamente")
    print("[OK] Las mejoras incluyen:")
    print("  - Detección de métricas implícitas en tablas simples")
    print("  - Tolerancia a typos en nombres de columnas")
    print("  - Exclusiones más inteligentes")
    print("  - Identificación de métricas calculadas esperadas")

if __name__ == "__main__":
    main()