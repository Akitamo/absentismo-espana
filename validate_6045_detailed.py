"""
Validador detallado tabla 6045: Nacional + Secciones CNAE (sin jornada)
Comparación completa INE CSV vs DuckDB
"""

import sys
from pathlib import Path
import pandas as pd
import duckdb
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración
sys.path.append(str(Path(__file__).parent))

def read_and_prepare_ine_csv(file_path):
    """
    Lee y prepara el CSV del INE tabla 6045
    """
    print("=" * 80)
    print("LEYENDO DATOS DEL INE - TABLA 6045")
    print("=" * 80)
    
    # Leer CSV con configuración correcta
    df = pd.read_csv(file_path, sep='\t', encoding='latin-1', decimal=',')
    
    print(f"\nArchivo: {file_path.name}")
    print(f"Columnas: {df.columns.tolist()}")
    print(f"Total registros: {len(df)}")
    
    # Filtrar solo 2025T1
    df_2025 = df[df['Periodo'] == '2025T1'].copy()
    print(f"Registros 2025T1: {len(df_2025)}")
    
    # Renombrar columna Total a Valor para consistencia
    df_2025.rename(columns={'Total': 'Valor'}, inplace=True)
    
    # Convertir valores a float
    df_2025['Valor'] = pd.to_numeric(df_2025['Valor'], errors='coerce')
    
    # Identificar columna de secciones
    seccion_col = None
    for col in df_2025.columns:
        if 'Secciones' in col or 'CNAE' in col:
            seccion_col = col
            break
    
    if seccion_col:
        print(f"\nColumna de secciones: {seccion_col}")
        print("\nAlgunas secciones CNAE encontradas:")
        secciones = df_2025[seccion_col].unique()
        for i, seccion in enumerate(secciones[:5]):
            print(f"  - {seccion[:60]}...")
        if len(secciones) > 5:
            print(f"  ... y {len(secciones)-5} más")
    
    return df_2025, seccion_col

def get_duckdb_data_detailed(db_path):
    """
    Extrae datos detallados de DuckDB para tabla 6045
    """
    print("\n" + "=" * 80)
    print("EXTRAYENDO DATOS DE DUCKDB - TABLA 6045")
    print("=" * 80)
    
    conn = duckdb.connect(str(db_path))
    
    # Query completa para obtener todos los datos de 2025T1 tabla 6045
    # NOTA: tabla 6045 NO tiene tipo_jornada (es NULL)
    query = """
    SELECT 
        tipo_jornada,
        cnae_nivel,
        cnae_codigo,
        cnae_nombre,
        metrica,
        causa,
        ROUND(valor, 1) as valor
    FROM observaciones_tiempo_trabajo
    WHERE fuente_tabla = '6045'
      AND periodo = '2025T1'
      AND ambito_territorial = 'NAC'
    ORDER BY cnae_nivel, cnae_codigo, metrica, causa
    """
    
    df = conn.execute(query).fetchdf()
    conn.close()
    
    print(f"\nTotal registros extraídos: {len(df)}")
    
    # Verificar tipo_jornada (debe ser NULL para 6045)
    jornadas_unicas = df['tipo_jornada'].unique()
    print(f"\nTipo jornada en BD: {jornadas_unicas}")
    if len(jornadas_unicas) == 1 and pd.isna(jornadas_unicas[0]):
        print("  Correcto: tipo_jornada es NULL (tabla sin jornada)")
    else:
        print("  Advertencia: se esperaba tipo_jornada = NULL")
    
    print("\nResumen por nivel CNAE:")
    for nivel in df['cnae_nivel'].unique():
        count = len(df[df['cnae_nivel'] == nivel])
        print(f"  - {nivel}: {count} registros")
    
    print("\nResumen por métrica:")
    for metrica in df['metrica'].unique():
        count = len(df[df['metrica'] == metrica])
        print(f"  - {metrica}: {count} registros")
    
    # Mostrar secciones CNAE
    secciones_db = df[df['cnae_nivel'] == 'SECCION']['cnae_codigo'].unique()
    secciones_db = [s for s in secciones_db if s is not None]
    print(f"\nSecciones CNAE en BD: {sorted(secciones_db)}")
    
    return df

def create_detailed_comparison(df_ine, df_db, seccion_col):
    """
    Crea comparación detallada entre INE y DuckDB para tabla 6045
    """
    print("\n" + "=" * 80)
    print("REALIZANDO COMPARACIÓN DETALLADA")
    print("=" * 80)
    
    comparisons = []
    
    # Mapeo de secciones CNAE (igual que tabla 6043, incluyendo correcciones para G y O)
    seccion_map = {
        'B_S Industria, construcción y servicios (excepto actividades de los hogares como empleadores y de organizaciones y organismos extraterritoriales)': ('TOTAL', None),
        'Total': ('TOTAL', None),
        'B. Industrias extractivas': ('SECCION', 'B'),
        'B Industrias extractivas': ('SECCION', 'B'),
        'C. Industria manufacturera': ('SECCION', 'C'),
        'C Industria manufacturera': ('SECCION', 'C'),
        'D. Suministro de energía eléctrica, gas, vapor y aire acondicionado': ('SECCION', 'D'),
        'D Suministro de energía eléctrica, gas, vapor y aire acondicionado': ('SECCION', 'D'),
        'E. Suministro de agua, actividades de saneamiento, gestión de residuos y descontaminación': ('SECCION', 'E'),
        'E Suministro de agua, actividades de saneamiento, gestión de residuos y descontaminación': ('SECCION', 'E'),
        'F. Construcción': ('SECCION', 'F'),
        'F Construcción': ('SECCION', 'F'),
        'G. Comercio al por mayor y al por menor; reparación de vehículos de motor y motocicletas': ('SECCION', 'G'),
        'G Comercio al por mayor y al por menor; reparación de vehículos de motor y motocicletas': ('SECCION', 'G'),
        'G Comercio al por mayor y al por menor, reparación de vehículos de motor y motocicletas': ('SECCION', 'G'),
        'H. Transporte y almacenamiento': ('SECCION', 'H'),
        'H Transporte y almacenamiento': ('SECCION', 'H'),
        'I. Hostelería': ('SECCION', 'I'),
        'I Hostelería': ('SECCION', 'I'),
        'J. Información y comunicaciones': ('SECCION', 'J'),
        'J Información y comunicaciones': ('SECCION', 'J'),
        'K. Actividades financieras y de seguros': ('SECCION', 'K'),
        'K Actividades financieras y de seguros': ('SECCION', 'K'),
        'L. Actividades inmobiliarias': ('SECCION', 'L'),
        'L Actividades inmobiliarias': ('SECCION', 'L'),
        'M. Actividades profesionales, científicas y técnicas': ('SECCION', 'M'),
        'M Actividades profesionales, científicas y técnicas': ('SECCION', 'M'),
        'N. Actividades administrativas y servicios auxiliares': ('SECCION', 'N'),
        'N Actividades administrativas y servicios auxiliares': ('SECCION', 'N'),
        'O. Administración Pública y defensa; Seguridad Social obligatoria': ('SECCION', 'O'),
        'O Administración Pública y defensa; Seguridad Social obligatoria': ('SECCION', 'O'),
        'O Administración Pública y defensa, Seguridad Social obligatoria': ('SECCION', 'O'),
        'P. Educación': ('SECCION', 'P'),
        'P Educación': ('SECCION', 'P'),
        'Q. Actividades sanitarias y de servicios sociales': ('SECCION', 'Q'),
        'Q Actividades sanitarias y de servicios sociales': ('SECCION', 'Q'),
        'R. Actividades artísticas, recreativas y de entretenimiento': ('SECCION', 'R'),
        'R Actividades artísticas, recreativas y de entretenimiento': ('SECCION', 'R'),
        'S. Otros servicios': ('SECCION', 'S'),
        'S Otros servicios': ('SECCION', 'S')
    }
    
    metrica_map = {
        'Horas pactadas': 'horas_pactadas',
        'Horas pagadas': 'horas_pagadas',
        'Horas efectivas': 'horas_efectivas',
        'Horas efectivas de trabajo': 'horas_efectivas',
        'Horas extraordinarias': 'horas_extraordinarias',
        'Horas extras por trabajador': 'horas_extraordinarias',
        'Horas no trabajadas': 'horas_no_trabajadas'
    }
    
    # Procesar cada registro del INE
    total_ine = len(df_ine)
    procesados = 0
    no_encontrados = []
    
    for idx, row_ine in df_ine.iterrows():
        procesados += 1
        if procesados % 10 == 0:
            print(f"  Procesando: {procesados}/{total_ine}")
        
        seccion_ine = row_ine[seccion_col] if seccion_col else ''
        tiempo_ine = row_ine['Tiempo de trabajo']
        valor_ine = row_ine['Valor']
        
        # Mapear valores
        cnae_nivel = None
        cnae_codigo = None
        
        # Primero buscar mapeo exacto
        if seccion_ine in seccion_map:
            cnae_nivel, cnae_codigo = seccion_map[seccion_ine]
        else:
            # Si no hay mapeo exacto, intentar identificar por patrones
            if seccion_ine.startswith('B_S'):
                cnae_nivel, cnae_codigo = 'TOTAL', None
            else:
                # Intentar extraer letra de sección
                for letra in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
                    if seccion_ine.startswith(f'{letra}. ') or seccion_ine.startswith(f'{letra} '):
                        cnae_nivel = 'SECCION'
                        cnae_codigo = letra
                        break
        
        metrica_db = metrica_map.get(tiempo_ine, None)
        
        if metrica_db and cnae_nivel:
            # Buscar valor correspondiente en DuckDB
            # Para tabla 6045, tipo_jornada debe ser NULL
            mask = (
                (df_db['tipo_jornada'].isna()) &
                (df_db['cnae_nivel'] == cnae_nivel) &
                (df_db['metrica'] == metrica_db)
            )
            
            if cnae_codigo:
                mask = mask & (df_db['cnae_codigo'] == cnae_codigo)
            else:
                mask = mask & (df_db['cnae_codigo'].isna())
            
            # Para HNT, solo tomar el total (causa = NULL)
            if metrica_db == 'horas_no_trabajadas':
                mask = mask & (df_db['causa'].isna())
            
            df_match = df_db[mask]
            
            if len(df_match) > 0:
                valor_db = df_match.iloc[0]['valor']
                diferencia = abs(valor_ine - valor_db)
                pct_diff = (diferencia / valor_ine * 100) if valor_ine != 0 else 0
                
                if diferencia < 0.1:
                    estado = 'OK'
                elif diferencia < 1:
                    estado = 'ADVERTENCIA'
                else:
                    estado = 'ERROR'
                
                comparisons.append({
                    'Sección CNAE': seccion_ine[:50],
                    'Métrica': tiempo_ine,
                    'Valor INE': valor_ine,
                    'Valor DuckDB': valor_db,
                    'Diferencia': round(diferencia, 2),
                    '% Diff': round(pct_diff, 2),
                    'Estado': estado
                })
            else:
                no_encontrados.append((seccion_ine, tiempo_ine))
                comparisons.append({
                    'Sección CNAE': seccion_ine[:50],
                    'Métrica': tiempo_ine,
                    'Valor INE': valor_ine,
                    'Valor DuckDB': 'NO ENCONTRADO',
                    'Diferencia': '-',
                    '% Diff': '-',
                    'Estado': 'SIN DATOS'
                })
    
    df_comparison = pd.DataFrame(comparisons)
    
    # Mostrar resumen
    print(f"\nTotal comparaciones: {len(comparisons)}")
    print("\nResumen por estado:")
    for estado in df_comparison['Estado'].unique():
        count = len(df_comparison[df_comparison['Estado'] == estado])
        print(f"  - {estado}: {count}")
    
    if no_encontrados:
        print(f"\nRegistros no encontrados: {len(no_encontrados)}")
        for i, (seccion, metrica) in enumerate(no_encontrados[:3]):
            print(f"  - {seccion[:40]} / {metrica}")
        if len(no_encontrados) > 3:
            print(f"  ... y {len(no_encontrados)-3} más")
    
    return df_comparison

def create_validation_excel(df_ine, df_db, df_comparison, output_path):
    """
    Crea Excel detallado con todas las validaciones
    """
    print("\n" + "=" * 80)
    print("GENERANDO EXCEL DE VALIDACIÓN")
    print("=" * 80)
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    info_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Hoja 1: Comparación Detallada
    ws1 = wb.active
    ws1.title = "Comparacion"
    
    # Escribir datos de comparación
    for r_idx, row in enumerate(dataframe_to_rows(df_comparison, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:  # Headers
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Colorear según estado
                estado = ws1.cell(row=r_idx, column=7).value  # Columna Estado
                if estado == 'OK':
                    for col in range(1, 8):
                        ws1.cell(row=r_idx, column=col).fill = ok_fill
                elif estado == 'ERROR':
                    for col in range(1, 8):
                        ws1.cell(row=r_idx, column=col).fill = error_fill
                elif estado == 'ADVERTENCIA':
                    for col in range(1, 8):
                        ws1.cell(row=r_idx, column=col).fill = warning_fill
                elif estado == 'SIN DATOS':
                    for col in range(1, 8):
                        ws1.cell(row=r_idx, column=col).fill = info_fill
    
    # Ajustar anchos de columna
    column_widths = [50, 30, 12, 12, 12, 10, 12]
    for i, width in enumerate(column_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = width
    
    # Hoja 2: Resumen
    ws2 = wb.create_sheet("Resumen")
    
    # Título
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "VALIDACIÓN TABLA 6045 - PERIODO 2025T1"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    ws2['A3'] = "Secciones CNAE (SIN tipo de jornada)"
    ws2['A3'].font = Font(italic=True)
    
    # Estadísticas
    ws2['A5'] = "ESTADÍSTICAS GENERALES"
    ws2['A5'].font = Font(bold=True, size=12)
    
    stats = df_comparison['Estado'].value_counts()
    row = 6
    for estado, count in stats.items():
        ws2.cell(row=row, column=1, value=f"{estado}:")
        ws2.cell(row=row, column=2, value=count)
        
        # Colorear según estado
        if estado == 'OK':
            ws2.cell(row=row, column=2).fill = ok_fill
        elif estado == 'ERROR':
            ws2.cell(row=row, column=2).fill = error_fill
        elif estado == 'ADVERTENCIA':
            ws2.cell(row=row, column=2).fill = warning_fill
        
        row += 1
    
    row += 1
    ws2.cell(row=row, column=1, value="TOTAL:")
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=len(df_comparison))
    ws2.cell(row=row, column=2).font = Font(bold=True)
    
    # Porcentaje de éxito
    ok_count = len(df_comparison[df_comparison['Estado'] == 'OK'])
    success_rate = (ok_count / len(df_comparison) * 100) if len(df_comparison) > 0 else 0
    
    row += 2
    ws2.cell(row=row, column=1, value="Tasa de éxito:")
    ws2.cell(row=row, column=2, value=f"{success_rate:.1f}%")
    ws2.cell(row=row, column=2).font = Font(bold=True, size=12)
    
    if success_rate >= 95:
        ws2.cell(row=row, column=2).fill = ok_fill
    elif success_rate >= 80:
        ws2.cell(row=row, column=2).fill = warning_fill
    else:
        ws2.cell(row=row, column=2).fill = error_fill
    
    # Información adicional
    row += 2
    ws2.cell(row=row, column=1, value="DETALLES:")
    ws2.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws2.cell(row=row, column=1, value="Registros INE:")
    ws2.cell(row=row, column=2, value=len(df_ine))
    row += 1
    ws2.cell(row=row, column=1, value="Registros DuckDB:")
    ws2.cell(row=row, column=2, value=len(df_db))
    row += 1
    ws2.cell(row=row, column=1, value="Tipo jornada:")
    ws2.cell(row=row, column=2, value="NULL (sin jornada)")
    
    # Hoja 3: Datos INE Original
    ws3 = wb.create_sheet("INE_Original")
    for r_idx, row in enumerate(dataframe_to_rows(df_ine, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    # Hoja 4: Datos DuckDB
    ws4 = wb.create_sheet("DuckDB_Datos")
    for r_idx, row in enumerate(dataframe_to_rows(df_db, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    # Guardar
    wb.save(output_path)
    print(f"Excel guardado: {output_path}")
    
    return output_path

def main():
    """
    Función principal de validación tabla 6045
    """
    print("\n" + "=" * 80)
    print("VALIDACIÓN DETALLADA TABLA 6045")
    print("Nacional + Secciones CNAE (SIN tipo de jornada)")
    print("Comparación INE CSV vs DuckDB - Periodo 2025T1")
    print("=" * 80 + "\n")
    
    # Configuración de rutas
    base_dir = Path(__file__).parent
    csv_file = base_dir / "data" / "INE" / "6045.csv"
    db_path = base_dir / "data" / "analysis.db"
    output_file = base_dir / "data" / "INE" / "validacion_6045_detallada.xlsx"
    
    try:
        # 1. Leer y preparar datos del INE
        df_ine, seccion_col = read_and_prepare_ine_csv(csv_file)
        
        # 2. Obtener datos de DuckDB
        df_db = get_duckdb_data_detailed(db_path)
        
        # 3. Realizar comparación detallada
        df_comparison = create_detailed_comparison(df_ine, df_db, seccion_col)
        
        # 4. Crear Excel de validación
        output_path = create_validation_excel(df_ine, df_db, df_comparison, output_file)
        
        # 5. Mostrar resumen final
        print("\n" + "=" * 80)
        print("VALIDACIÓN COMPLETADA - TABLA 6045")
        print("=" * 80)
        
        # Calcular estadísticas finales
        ok_count = len(df_comparison[df_comparison['Estado'] == 'OK'])
        total_count = len(df_comparison)
        success_rate = (ok_count / total_count * 100) if total_count > 0 else 0
        
        print(f"\nResultados:")
        print(f"  - Comparaciones exitosas: {ok_count}/{total_count}")
        print(f"  - Tasa de éxito: {success_rate:.1f}%")
        
        if success_rate >= 95:
            print(f"\n[OK] Validación exitosa - Alta coincidencia con datos INE")
        elif success_rate >= 80:
            print(f"\n[ADVERTENCIA] Validación con observaciones - Revisar discrepancias")
        else:
            print(f"\n[ERROR] Validación con problemas - Requiere revisión detallada")
        
        print(f"\nArchivo Excel generado: {output_file.name}")
        print(f"Abre el archivo para revisar:")
        print(f"  1. Hoja 'Comparacion': Detalle línea por línea")
        print(f"  2. Hoja 'Resumen': Estadísticas de validación")
        print(f"  3. Hoja 'INE_Original': Datos fuente del INE")
        print(f"  4. Hoja 'DuckDB_Datos': Datos en nuestra BD")
        
        print(f"\nCaracterística especial tabla 6045:")
        print(f"  - NO tiene tipo de jornada (tipo_jornada = NULL)")
        print(f"  - Secciones CNAE detalladas (B-S)")
        
    except Exception as e:
        print(f"\nError durante la validación: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()