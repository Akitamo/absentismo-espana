#!/usr/bin/env python3
"""
Script para extraer TODAS las métricas detalladas de las 35 tablas del INE.
Genera un informe completo con nombres exactos de columnas métricas y su categorización.
"""

import pandas as pd
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set, Tuple
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de rutas
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data" / "raw" / "csv"
CONFIG_FILE = BASE_DIR / "config" / "tables.json"
REPORTS_DIR = BASE_DIR / "data" / "exploration_reports"

# Dimensiones conocidas del análisis previo
KNOWN_DIMENSIONS = {
    'periodo', 'Periodo', 'PERIODO',
    'Sectores', 'SECTORES', 'sectores',
    'Comunidades y Ciudades Autónomas', 'CCAA', 'Comunidades',
    'Tipo de jornada', 'TIPO_JORNADA', 'TipoJornada',
    'Tamaño del establecimiento', 'TAMAÑO', 'Tamaño',
    'Secciones', 'SECCIONES', 'Secciones CNAE',
    'Divisiones', 'DIVISIONES', 'Divisiones CNAE',
    'Total', 'TOTAL'
}

# Categorías de métricas según metodología INE
METRIC_CATEGORIES = {
    'COSTES_LABORALES': [
        'coste', 'cost', 'cotizacion', 'cotización', 'subvencion', 'subvención',
        'despido', 'prestacion', 'prestación', 'percepcion', 'percepción'
    ],
    'TIEMPO_TRABAJO': [
        'hora', 'tiempo', 'jornada', 'efectiva', 'pactada', 'pagada', 
        'extra', 'extraordinaria', 'no trabajada', 'i.t.', 'it', 'incapacidad'
    ],
    'COSTE_SALARIAL': [
        'salarial', 'salario', 'ordinario', 'extraordinario', 'atrasado'
    ],
    'VACANTES': [
        'vacante', 'puesto', 'plaza'
    ],
    'MOTIVOS_NO_VACANTES': [
        'motivo', 'razon', 'razón', 'causa'
    ],
    'INDICES': [
        'indice', 'índice', 'tasa', 'variacion', 'variación'
    ],
    'OTROS': []  # Para métricas que no encajan en las anteriores
}

def load_config() -> Dict:
    """Carga la configuración de tablas desde el archivo JSON."""
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Aplanar estructura de categorías a diccionario simple código->info
    tables = {}
    for categoria_key, categoria_data in data.get('categorias', {}).items():
        for table_code, table_info in categoria_data.get('tablas', {}).items():
            table_info['categoria'] = categoria_data.get('descripcion', categoria_key)
            tables[table_code] = table_info
    
    return tables

def detect_encoding(file_path: Path) -> str:
    """Detecta la codificación del archivo CSV."""
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                f.read(1000)
            return encoding
        except UnicodeDecodeError:
            continue
    return 'utf-8'

def is_numeric_column(series: pd.Series) -> bool:
    """Determina si una columna contiene datos numéricos."""
    if series.dtype in [np.float64, np.int64]:
        return True
    
    # Intentar convertir a numérico
    try:
        # Limpiar valores y convertir
        cleaned = series.str.replace(',', '.').str.replace(' ', '')
        pd.to_numeric(cleaned, errors='raise')
        return True
    except:
        return False

def categorize_metric(metric_name: str) -> str:
    """Categoriza una métrica según su nombre."""
    metric_lower = metric_name.lower()
    
    for category, keywords in METRIC_CATEGORIES.items():
        if category == 'OTROS':
            continue
        for keyword in keywords:
            if keyword in metric_lower:
                return category
    
    return 'OTROS'

def extract_table_metrics(table_code: str, table_info: Dict) -> Dict:
    """Extrae todas las métricas de una tabla específica."""
    # Buscar archivo CSV
    csv_pattern = f"{table_code}_*.csv"
    csv_files = list(DATA_DIR.glob(csv_pattern))
    
    if not csv_files:
        print(f"  [!] No se encontró archivo para tabla {table_code}")
        return None
    
    csv_file = csv_files[0]
    print(f"  Analizando: {csv_file.name}")
    
    # Detectar encoding y leer archivo
    encoding = detect_encoding(csv_file)
    
    try:
        # Leer CSV con diferentes separadores posibles
        df = None
        for sep in [';', ',', '\t']:
            try:
                df = pd.read_csv(csv_file, encoding=encoding, sep=sep, decimal=',')
                if len(df.columns) > 1:  # Verificar que se separó correctamente
                    break
            except:
                continue
        
        if df is None or len(df) == 0:
            print(f"  [!] No se pudo leer el archivo {csv_file.name}")
            return None
        
        # Identificar estructura del CSV (formato largo vs ancho)
        metrics = []
        dimensions = []
        
        # Buscar columnas que contienen las métricas (formato largo)
        metric_columns = []
        for col in df.columns:
            col_lower = col.lower()
            # Columnas que típicamente contienen nombres de métricas en formato largo
            if any(term in col_lower for term in ['componente', 'concepto', 'variable', 'indicador', 
                                                   'tipo', 'motivo', 'causa']):
                metric_columns.append(col)
        
        # Si encontramos columnas de métricas, extraer valores únicos
        if metric_columns:
            for metric_col in metric_columns:
                unique_metrics = df[metric_col].dropna().unique()
                for metric_name in unique_metrics:
                    if metric_name and str(metric_name).strip() and str(metric_name) != 'nan':
                        metric_clean = str(metric_name).strip()
                        # Verificar que no sea una dimensión
                        if not any(dim.lower() in metric_clean.lower() for dim in ['total', 'todos']):
                            metrics.append({
                                'nombre': metric_clean,
                                'categoria': categorize_metric(metric_clean),
                                'columna_origen': metric_col,
                                'tipo_dato': 'valor_categorico',
                                'valores_unicos': 1,
                                'tiene_nulos': False,
                                'valores_ejemplo': []
                            })
        
        # Identificar dimensiones
        for col in df.columns:
            col_clean = col.strip()
            col_lower = col_clean.lower()
            
            # Saltar columna 'Total' que contiene valores
            if col_lower == 'total':
                continue
            
            # Verificar si es dimensión conocida
            is_dimension = False
            if any(term in col_lower for term in ['periodo', 'sector', 'actividad', 'comunidad', 
                                                   'ccaa', 'autonoma', 'jornada', 'tamaño', 
                                                   'establecimiento', 'seccion', 'division', 
                                                   'cnae']):
                dimensions.append(col_clean)
                is_dimension = True
            
            # Si no es dimensión conocida ni columna de métricas, verificar si es numérica
            if not is_dimension and col not in metric_columns:
                if col_lower == 'total' or 'valor' in col_lower:
                    # Esta es la columna con valores numéricos
                    continue
                elif is_numeric_column(df[col]):
                    # Es una métrica en formato ancho (columna = métrica)
                    unique_count = df[col].nunique()
                    if unique_count > 20:  # Probablemente sea una métrica real
                        metrics.append({
                            'nombre': col_clean,
                            'categoria': categorize_metric(col_clean),
                            'tipo_dato': str(df[col].dtype),
                            'valores_unicos': unique_count,
                            'tiene_nulos': df[col].isnull().any(),
                            'valores_ejemplo': df[col].dropna().head(3).tolist() if len(df[col].dropna()) > 0 else []
                        })
        
        # Eliminar duplicados en métricas
        metrics_unique = []
        seen_metrics = set()
        for m in metrics:
            if m['nombre'] not in seen_metrics:
                metrics_unique.append(m)
                seen_metrics.add(m['nombre'])
        
        # Crear resultado estructurado
        result = {
            'codigo': table_code,
            'nombre': table_info.get('nombre', ''),
            'categoria': table_info.get('categoria', ''),
            'archivo': csv_file.name,
            'total_columnas': len(df.columns),
            'total_filas': len(df),
            'dimensiones': dimensions,
            'total_dimensiones': len(dimensions),
            'metricas': metrics_unique,
            'total_metricas': len(metrics_unique),
            'metricas_por_categoria': {},
            'formato_datos': 'largo' if metric_columns else 'ancho'
        }
        
        # Agrupar métricas por categoría
        for categoria in METRIC_CATEGORIES.keys():
            metricas_categoria = [m for m in metrics_unique if m['categoria'] == categoria]
            if metricas_categoria:
                result['metricas_por_categoria'][categoria] = {
                    'total': len(metricas_categoria),
                    'metricas': sorted([m['nombre'] for m in metricas_categoria])
                }
        
        return result
        
    except Exception as e:
        print(f"  [!] Error procesando {csv_file.name}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generate_excel_report(all_metrics: Dict, output_file: Path):
    """Genera un informe Excel con todas las métricas detalladas."""
    wb = Workbook()
    
    # Hoja 1: Resumen general
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados resumen
    headers_resumen = ["Código", "Nombre Tabla", "Categoría", "Total Métricas", 
                      "Costes Lab.", "Tiempo Trab.", "Coste Sal.", "Vacantes", 
                      "Motivos NV", "Índices", "Otros", "Dimensiones"]
    
    for col, header in enumerate(headers_resumen, 1):
        cell = ws_resumen.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos resumen
    row = 2
    for table_code, data in sorted(all_metrics.items()):
        if data:
            ws_resumen.cell(row=row, column=1, value=table_code).border = border
            ws_resumen.cell(row=row, column=2, value=data['nombre']).border = border
            ws_resumen.cell(row=row, column=3, value=data['categoria']).border = border
            ws_resumen.cell(row=row, column=4, value=data['total_metricas']).border = border
            
            col = 5
            for cat in ['COSTES_LABORALES', 'TIEMPO_TRABAJO', 'COSTE_SALARIAL', 
                       'VACANTES', 'MOTIVOS_NO_VACANTES', 'INDICES', 'OTROS']:
                valor = data['metricas_por_categoria'].get(cat, {}).get('total', 0)
                ws_resumen.cell(row=row, column=col, value=valor).border = border
                col += 1
            
            ws_resumen.cell(row=row, column=12, value=', '.join(data['dimensiones'][:3])).border = border
            row += 1
    
    # Ajustar anchos de columna
    for col in range(1, 13):
        ws_resumen.column_dimensions[get_column_letter(col)].width = 15
    
    # Hoja 2: Detalle de métricas por tabla
    ws_detalle = wb.create_sheet("Detalle Métricas")
    
    headers_detalle = ["Código Tabla", "Nombre Tabla", "Métrica", "Categoría", 
                       "Tipo Dato", "Valores Únicos", "Tiene Nulos", "Ejemplos"]
    
    for col, header in enumerate(headers_detalle, 1):
        cell = ws_detalle.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 2
    for table_code, data in sorted(all_metrics.items()):
        if data and data['metricas']:
            for metrica in data['metricas']:
                ws_detalle.cell(row=row, column=1, value=table_code).border = border
                ws_detalle.cell(row=row, column=2, value=data['nombre']).border = border
                ws_detalle.cell(row=row, column=3, value=metrica['nombre']).border = border
                ws_detalle.cell(row=row, column=4, value=metrica['categoria']).border = border
                ws_detalle.cell(row=row, column=5, value=metrica['tipo_dato']).border = border
                ws_detalle.cell(row=row, column=6, value=metrica['valores_unicos']).border = border
                ws_detalle.cell(row=row, column=7, value="Sí" if metrica['tiene_nulos'] else "No").border = border
                
                # Ejemplos
                ejemplos = ', '.join([str(v) for v in metrica['valores_ejemplo'][:3]])
                ws_detalle.cell(row=row, column=8, value=ejemplos).border = border
                row += 1
    
    # Ajustar anchos
    ws_detalle.column_dimensions['A'].width = 12
    ws_detalle.column_dimensions['B'].width = 40
    ws_detalle.column_dimensions['C'].width = 50
    ws_detalle.column_dimensions['D'].width = 20
    ws_detalle.column_dimensions['E'].width = 15
    ws_detalle.column_dimensions['F'].width = 15
    ws_detalle.column_dimensions['G'].width = 12
    ws_detalle.column_dimensions['H'].width = 30
    
    # Hoja 3: Matriz de métricas
    ws_matriz = wb.create_sheet("Matriz Métricas")
    
    # Obtener todas las métricas únicas
    all_metric_names = set()
    for data in all_metrics.values():
        if data:
            for metrica in data['metricas']:
                all_metric_names.add(metrica['nombre'])
    
    metric_names_sorted = sorted(list(all_metric_names))
    
    # Encabezados: Tabla + todas las métricas
    ws_matriz.cell(row=1, column=1, value="Tabla")
    ws_matriz.cell(row=1, column=1).fill = header_fill
    ws_matriz.cell(row=1, column=1).font = header_font
    
    for col, metric_name in enumerate(metric_names_sorted, 2):
        cell = ws_matriz.cell(row=1, column=col, value=metric_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(text_rotation=90, horizontal='center')
    
    # Llenar matriz
    row = 2
    for table_code, data in sorted(all_metrics.items()):
        if data:
            ws_matriz.cell(row=row, column=1, value=f"{table_code} - {data['nombre'][:30]}")
            
            # Marcar qué métricas tiene esta tabla
            table_metrics = {m['nombre'] for m in data['metricas']}
            for col, metric_name in enumerate(metric_names_sorted, 2):
                if metric_name in table_metrics:
                    cell = ws_matriz.cell(row=row, column=col, value="X")
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            row += 1
    
    # Ajustar anchos
    ws_matriz.column_dimensions['A'].width = 50
    for col in range(2, len(metric_names_sorted) + 2):
        ws_matriz.column_dimensions[get_column_letter(col)].width = 3
    
    # Guardar archivo
    wb.save(output_file)
    print(f"\n[OK] Informe Excel guardado en: {output_file}")

def main():
    """Función principal."""
    print("=" * 80)
    print("EXTRACCIÓN EXHAUSTIVA DE MÉTRICAS - TODAS LAS TABLAS INE")
    print("=" * 80)
    
    # Cargar configuración
    config = load_config()
    
    # Procesar todas las tablas
    all_metrics = {}
    total_processed = 0
    total_metrics_found = 0
    
    for table_code, table_info in config.items():
        print(f"\nProcesando tabla {table_code}: {table_info['nombre']}")
        
        result = extract_table_metrics(table_code, table_info)
        
        if result:
            all_metrics[table_code] = result
            total_processed += 1
            total_metrics_found += result['total_metricas']
            print(f"  -> {result['total_metricas']} métricas encontradas")
            print(f"  -> {result['total_dimensiones']} dimensiones identificadas")
    
    # Generar timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Guardar JSON detallado
    json_file = REPORTS_DIR / f"metricas_exhaustivas_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(all_metrics, f, ensure_ascii=False, indent=2)
    print(f"\n[OK] Datos JSON guardados en: {json_file}")
    
    # Generar Excel
    excel_file = REPORTS_DIR / f"metricas_exhaustivas_{timestamp}.xlsx"
    generate_excel_report(all_metrics, excel_file)
    
    # Resumen final
    print("\n" + "=" * 80)
    print("RESUMEN FINAL")
    print("=" * 80)
    print(f"Tablas procesadas: {total_processed}/{len(config)}")
    print(f"Total métricas encontradas: {total_metrics_found}")
    
    # Resumen por categoría
    category_totals = {cat: 0 for cat in METRIC_CATEGORIES.keys()}
    for data in all_metrics.values():
        if data:
            for cat, info in data['metricas_por_categoria'].items():
                category_totals[cat] += info['total']
    
    print("\nMétricas por categoría:")
    for cat, total in category_totals.items():
        if total > 0:
            print(f"  - {cat}: {total}")
    
    print("\n[OK] Proceso completado exitosamente")

if __name__ == "__main__":
    main()