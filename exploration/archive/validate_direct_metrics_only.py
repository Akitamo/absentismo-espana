#!/usr/bin/env python3
"""
Script para validar SOLO las métricas DIRECTAS extraídas contra la metodología oficial del INE.
Excluye métricas calculadas y las no relevantes para análisis de absentismo.
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configuración de rutas
BASE_DIR = Path(__file__).parent.parent
REPORTS_DIR = BASE_DIR / "data" / "exploration_reports"

# MÉTRICAS DIRECTAS DEL INE (excluyendo calculadas y no relevantes)
METRICAS_DIRECTAS_INE = {
    'COSTES_LABORALES': [
        'Coste total',  # Posiblemente sinónimo de Coste laboral total
        'Coste laboral total',
        'Coste salarial total',
        'Coste salarial ordinario',
        'Coste salarial pagos extraordinarios',
        'Coste salarial pagos atrasados',
        'Coste salarial extraordinario',
        'Otros costes',
        'Coste por percepciones no salariales',
        'Coste por I.T.',
        'Coste por desempleo',
        'Coste por otras prestaciones sociales directas',
        'Coste por otras percepciones no salariales',
        'Coste por cotizaciones obligatorias',
        'Coste por contingencias comunes',
        'Coste por desempleo, Fogasa y F. Profesional',
        'Coste por otras cotizaciones sociales obligatorias',
        'Coste por despido',
        # 'Subvenciones y bonificaciones de la Seguridad Social', # EXCLUIDA - No relevante para absentismo
    ],
    'TIEMPO_TRABAJO': [
        'Horas pactadas',
        'Horas pagadas',
        'Horas efectivas',
        'Horas no trabajadas',
        'Horas no trabajadas por vacaciones',
        'Horas no trabajadas por fiestas',
        'Horas no trabajadas por I.T.',
        'Horas no trabajadas por maternidad',
        'Horas no trabajadas por permisos remunerados',
        'Horas no trabajadas por razones técnicas',
        'Horas no trabajadas por conflictos laborales',
        'Horas extras',
        'Horas extraordinarias'  # Sinónimo de Horas extras
    ],
    'VACANTES': [
        'Número de vacantes',
        'Motivos por los que no existen vacantes'  # Incluye múltiples sub-métricas
    ]
}

# MÉTRICAS CALCULADAS (NO están directamente en CSVs)
METRICAS_CALCULADAS = {
    'Percepciones por día de I.T.',  # Calculada a partir de otros datos
    'Coste indemnización trabajador despedido',  # Calculada
    'Coste por hora extra'  # Calculada
}

# Sub-métricas de "Motivos por los que no existen vacantes"
MOTIVOS_NO_VACANTES = {
    'Elevado coste de contratación',
    'No se necesita ningún trabajador',
    'Otros motivos'  # Puede aparecer en algunos CSVs
}

# Equivalencias y sinónimos mejorados
EQUIVALENCIAS = {
    # Costes
    'Coste laboral total': ['Coste total por trabajador', 'Coste laboral', 'Coste total'],
    'Coste por I.T.': ['Coste I.T.', 'Pagos por Incapacidad Temporal', 'Coste por incapacidad temporal'],
    'Coste por desempleo': ['Coste por desempleo parcial', 'Pagos por desempleo'],
    'Coste por despido': ['Indemnizaciones por despido', 'Coste indemnización'],
    'Coste por contingencias comunes': ['Contingencias comunes'],
    'Coste por desempleo, Fogasa y F. Profesional': [
        'Desempleo, Fogasa y Formación Profesional',
        'Coste por desempleo,  Fogasa y F. Profesional'  # Con doble espacio
    ],
    
    # Tiempo
    'Horas efectivas': ['Horas realmente trabajadas'],
    'Horas no trabajadas': ['Tiempo no trabajado'],
    'Horas extras': ['Horas extraordinarias', 'Horas extras por trabajador'],
    'Horas extraordinarias': ['Horas extras', 'Horas extras por trabajador'],
    'Horas no trabajadas por I.T.': ['Horas no trabajadas por I.T', 'Horas no trabajadas por incapacidad temporal'],
    'Horas no trabajadas por vacaciones': ['Horas no trabajadas por vacaciones y fiestas'],
    'Horas no trabajadas por razones técnicas': ['Horas no trabajadas por razones técnicas o económicas'],
    'Horas no trabajadas por conflictos laborales': ['Horas no trabajadas por conflictividad laboral'],
    
    # Vacantes
    'Número de vacantes': ['Vacantes', 'Puestos vacantes'],
    'Motivos por los que no existen vacantes': [
        'Motivos no vacantes', 
        'Razones de no vacantes',
        'Elevado coste de contratación',
        'No se necesita ningún trabajador'
    ]
}

def normalize_metric_name(name: str) -> str:
    """Normaliza un nombre de métrica para comparación."""
    # Convertir a minúsculas y eliminar espacios extra
    normalized = name.lower().strip()
    # Eliminar caracteres especiales comunes
    normalized = normalized.replace('.', '').replace(',', '')
    # Normalizar espacios
    normalized = ' '.join(normalized.split())
    # Normalizar I.T. vs IT
    normalized = normalized.replace('it', 'it')  # Mantener consistencia
    return normalized

def is_sub_metric_of_motivos(metric: str) -> bool:
    """Verifica si una métrica es un sub-tipo de 'Motivos por los que no existen vacantes'."""
    metric_norm = normalize_metric_name(metric)
    motivos_norm = {normalize_metric_name(m) for m in MOTIVOS_NO_VACANTES}
    
    # Verificar si es uno de los motivos conocidos
    if metric_norm in motivos_norm:
        return True
    
    # Verificar patrones comunes
    if 'elevado coste' in metric_norm or 'no se necesita' in metric_norm:
        return True
    
    return False

def find_metric_in_official(metric: str, official_metrics: Dict[str, List[str]]) -> tuple:
    """
    Busca una métrica en el listado oficial.
    Retorna (categoría, métrica_oficial, coincidencia_exacta)
    """
    metric_norm = normalize_metric_name(metric)
    
    # Primero verificar si es un sub-tipo de motivos de no vacantes
    if is_sub_metric_of_motivos(metric):
        return 'VACANTES', 'Motivos por los que no existen vacantes', True
    
    # Buscar coincidencia exacta o equivalente
    for category, metrics_list in official_metrics.items():
        for official_metric in metrics_list:
            official_norm = normalize_metric_name(official_metric)
            
            # Coincidencia exacta
            if metric_norm == official_norm:
                return category, official_metric, True
            
            # Buscar en equivalencias
            if official_metric in EQUIVALENCIAS:
                for equiv in EQUIVALENCIAS[official_metric]:
                    if metric_norm == normalize_metric_name(equiv):
                        return category, official_metric, True
            
            # Coincidencia parcial (una contenida en la otra)
            if metric_norm in official_norm or official_norm in metric_norm:
                # Evitar falsos positivos con "horas no trabajadas"
                if 'horas no trabajadas' in official_norm and 'horas no trabajadas' in metric_norm:
                    # Es una variación específica de horas no trabajadas
                    return category, official_metric, True
                return category, official_metric, False
    
    return None, None, False

def load_extracted_metrics(json_file: Path = None) -> Dict:
    """Carga las métricas extraídas del archivo JSON más reciente."""
    if not json_file:
        # Buscar archivo de métricas más reciente (primero enhanced, luego limpias)
        json_files = list(REPORTS_DIR.glob("metricas_enhanced_*.json"))
        if not json_files:
            json_files = list(REPORTS_DIR.glob("metricas_limpias_*.json"))
        if not json_files:
            raise FileNotFoundError("No se encontró archivo de métricas extraídas")
        json_file = max(json_files, key=lambda x: x.stat().st_mtime)
    
    with open(json_file, 'r', encoding='utf-8') as f:
        return json.load(f)

def validate_direct_metrics(extracted_data: Dict) -> Dict:
    """Valida solo las métricas DIRECTAS extraídas contra las oficiales del INE."""
    
    # Recopilar todas las métricas únicas extraídas
    extracted_metrics = set()
    for table_data in extracted_data.values():
        if table_data and 'metricas' in table_data:
            for metric in table_data['metricas']:
                extracted_metrics.add(metric['nombre'])
    
    # Validar cada métrica extraída
    validation_results = {
        'metricas_validadas': [],
        'metricas_no_encontradas': [],
        'metricas_parciales': [],
        'metricas_calculadas_detectadas': [],
        'resumen_por_categoria': {}
    }
    
    for metric in sorted(extracted_metrics):
        # Verificar si es una métrica calculada
        if metric in METRICAS_CALCULADAS:
            validation_results['metricas_calculadas_detectadas'].append(metric)
            continue
        
        category, official, exact = find_metric_in_official(metric, METRICAS_DIRECTAS_INE)
        
        if category and exact:
            validation_results['metricas_validadas'].append({
                'metrica_extraida': metric,
                'metrica_oficial': official,
                'categoria': category
            })
        elif category and not exact:
            validation_results['metricas_parciales'].append({
                'metrica_extraida': metric,
                'posible_oficial': official,
                'categoria': category
            })
        else:
            # Verificar una vez más si podría ser calculada
            if any(calc_term in metric.lower() for calc_term in ['día', 'indemnización', 'hora extra']):
                validation_results['metricas_calculadas_detectadas'].append(metric)
            else:
                validation_results['metricas_no_encontradas'].append(metric)
    
    # Buscar métricas oficiales directas no extraídas
    all_official = []
    for cat, metrics in METRICAS_DIRECTAS_INE.items():
        all_official.extend([(m, cat) for m in metrics])
    
    extracted_normalized = {normalize_metric_name(m) for m in extracted_metrics}
    
    validation_results['metricas_oficiales_faltantes'] = []
    for official_metric, category in all_official:
        found = False
        official_norm = normalize_metric_name(official_metric)
        
        # Verificar si está extraída
        if official_norm in extracted_normalized:
            found = True
        else:
            # Verificar equivalencias
            if official_metric in EQUIVALENCIAS:
                for equiv in EQUIVALENCIAS[official_metric]:
                    if normalize_metric_name(equiv) in extracted_normalized:
                        found = True
                        break
            
            # Para motivos de no vacantes, verificar sub-métricas
            if official_metric == 'Motivos por los que no existen vacantes':
                for extracted in extracted_metrics:
                    if is_sub_metric_of_motivos(extracted):
                        found = True
                        break
        
        if not found:
            # Verificar coincidencias parciales
            for extracted in extracted_normalized:
                if official_norm in extracted or extracted in official_norm:
                    found = True
                    break
        
        if not found:
            validation_results['metricas_oficiales_faltantes'].append({
                'metrica': official_metric,
                'categoria': category
            })
    
    # Calcular resumen
    for cat in METRICAS_DIRECTAS_INE.keys():
        total_oficiales = len(METRICAS_DIRECTAS_INE[cat])
        encontradas = len([m for m in validation_results['metricas_validadas'] 
                          if m['categoria'] == cat])
        parciales = len([m for m in validation_results['metricas_parciales'] 
                        if m['categoria'] == cat])
        
        validation_results['resumen_por_categoria'][cat] = {
            'total_oficiales': total_oficiales,
            'encontradas_exactas': encontradas,
            'encontradas_parciales': parciales,
            'porcentaje_cobertura': round((encontradas / total_oficiales) * 100, 1) if total_oficiales > 0 else 0
        }
    
    # Resumen global
    total_oficiales = sum(len(m) for m in METRICAS_DIRECTAS_INE.values())
    total_validadas = len(validation_results['metricas_validadas'])
    total_parciales = len(validation_results['metricas_parciales'])
    total_calculadas = len(validation_results['metricas_calculadas_detectadas'])
    
    validation_results['resumen_global'] = {
        'total_metricas_oficiales_directas': total_oficiales,
        'total_metricas_extraidas': len(extracted_metrics),
        'metricas_validadas': total_validadas,
        'metricas_parciales': total_parciales,
        'metricas_calculadas_detectadas': total_calculadas,
        'metricas_no_encontradas': len(validation_results['metricas_no_encontradas']),
        'porcentaje_precision': round((total_validadas / (len(extracted_metrics) - total_calculadas)) * 100, 1) if (len(extracted_metrics) - total_calculadas) > 0 else 0,
        'porcentaje_cobertura': round((total_validadas / total_oficiales) * 100, 1) if total_oficiales > 0 else 0
    }
    
    return validation_results

def generate_validation_report(validation: Dict, output_file: Path):
    """Genera un informe Excel de validación de métricas directas."""
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    calc_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Hoja 1: Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Validación Directas"
    
    # Título
    ws_resumen.merge_cells('A1:D1')
    ws_resumen['A1'] = "VALIDACIÓN DE MÉTRICAS DIRECTAS (NO CALCULADAS)"
    ws_resumen['A1'].font = Font(size=14, bold=True)
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    ws_resumen.merge_cells('A2:D2')
    ws_resumen['A2'] = "Excluye: Métricas calculadas y Subvenciones (no relevante para absentismo)"
    ws_resumen['A2'].font = Font(size=10, italic=True)
    ws_resumen['A2'].alignment = Alignment(horizontal='center')
    
    # Resumen global
    row = 4
    ws_resumen[f'A{row}'] = "RESUMEN GLOBAL"
    ws_resumen[f'A{row}'].font = Font(bold=True)
    row += 1
    
    global_data = validation['resumen_global']
    ws_resumen[f'A{row}'] = "Total métricas oficiales DIRECTAS:"
    ws_resumen[f'B{row}'] = global_data['total_metricas_oficiales_directas']
    row += 1
    ws_resumen[f'A{row}'] = "Total métricas extraídas:"
    ws_resumen[f'B{row}'] = global_data['total_metricas_extraidas']
    row += 1
    ws_resumen[f'A{row}'] = "Métricas validadas (exactas):"
    ws_resumen[f'B{row}'] = global_data['metricas_validadas']
    ws_resumen[f'B{row}'].fill = success_fill
    row += 1
    ws_resumen[f'A{row}'] = "Métricas parciales:"
    ws_resumen[f'B{row}'] = global_data['metricas_parciales']
    ws_resumen[f'B{row}'].fill = warning_fill
    row += 1
    ws_resumen[f'A{row}'] = "Métricas calculadas detectadas:"
    ws_resumen[f'B{row}'] = global_data['metricas_calculadas_detectadas']
    ws_resumen[f'B{row}'].fill = calc_fill
    row += 1
    ws_resumen[f'A{row}'] = "Métricas no identificadas:"
    ws_resumen[f'B{row}'] = global_data['metricas_no_encontradas']
    ws_resumen[f'B{row}'].fill = error_fill if global_data['metricas_no_encontradas'] > 5 else warning_fill
    row += 2
    
    ws_resumen[f'A{row}'] = "% Precisión (validadas/extraídas no calculadas):"
    ws_resumen[f'B{row}'] = f"{global_data['porcentaje_precision']}%"
    row += 1
    ws_resumen[f'A{row}'] = "% Cobertura (validadas/oficiales directas):"
    ws_resumen[f'B{row}'] = f"{global_data['porcentaje_cobertura']}%"
    if global_data['porcentaje_cobertura'] >= 80:
        ws_resumen[f'B{row}'].fill = success_fill
    
    # Resumen por categoría
    row += 3
    ws_resumen[f'A{row}'] = "RESUMEN POR CATEGORÍA"
    ws_resumen[f'A{row}'].font = Font(bold=True)
    row += 1
    
    headers = ["Categoría", "Total Oficial", "Encontradas", "Parciales", "% Cobertura"]
    for col, header in enumerate(headers, 1):
        cell = ws_resumen.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    for cat, data in validation['resumen_por_categoria'].items():
        ws_resumen.cell(row=row, column=1, value=cat)
        ws_resumen.cell(row=row, column=2, value=data['total_oficiales'])
        ws_resumen.cell(row=row, column=3, value=data['encontradas_exactas'])
        ws_resumen.cell(row=row, column=4, value=data['encontradas_parciales'])
        ws_resumen.cell(row=row, column=5, value=f"{data['porcentaje_cobertura']}%")
        
        # Colorear según cobertura
        if data['porcentaje_cobertura'] >= 80:
            ws_resumen.cell(row=row, column=5).fill = success_fill
        elif data['porcentaje_cobertura'] >= 60:
            ws_resumen.cell(row=row, column=5).fill = warning_fill
        else:
            ws_resumen.cell(row=row, column=5).fill = error_fill
        row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 45
    ws_resumen.column_dimensions['B'].width = 20
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 15
    ws_resumen.column_dimensions['E'].width = 15
    
    # Hoja 2: Métricas Validadas
    ws_validadas = wb.create_sheet("Métricas Directas Validadas")
    headers = ["Métrica Extraída", "Métrica Oficial INE", "Categoría"]
    for col, header in enumerate(headers, 1):
        cell = ws_validadas.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 2
    for item in validation['metricas_validadas']:
        ws_validadas.cell(row=row, column=1, value=item['metrica_extraida'])
        ws_validadas.cell(row=row, column=2, value=item['metrica_oficial'])
        ws_validadas.cell(row=row, column=3, value=item['categoria'])
        row += 1
    
    ws_validadas.column_dimensions['A'].width = 50
    ws_validadas.column_dimensions['B'].width = 50
    ws_validadas.column_dimensions['C'].width = 25
    
    # Hoja 3: Métricas Calculadas Detectadas
    if validation['metricas_calculadas_detectadas']:
        ws_calculadas = wb.create_sheet("Métricas Calculadas")
        ws_calculadas['A1'] = "Métricas Calculadas (No son directas de CSVs)"
        ws_calculadas['A1'].fill = header_fill
        ws_calculadas['A1'].font = header_font
        
        row = 2
        for metric in validation['metricas_calculadas_detectadas']:
            ws_calculadas.cell(row=row, column=1, value=metric)
            ws_calculadas.cell(row=row, column=1).fill = calc_fill
            row += 1
        
        ws_calculadas.column_dimensions['A'].width = 60
    
    # Hoja 4: Métricas Directas Faltantes
    if validation['metricas_oficiales_faltantes']:
        ws_faltantes = wb.create_sheet("Métricas Directas Faltantes")
        headers = ["Métrica Oficial INE (Directa)", "Categoría"]
        for col, header in enumerate(headers, 1):
            cell = ws_faltantes.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for item in validation['metricas_oficiales_faltantes']:
            ws_faltantes.cell(row=row, column=1, value=item['metrica'])
            ws_faltantes.cell(row=row, column=2, value=item['categoria'])
            ws_faltantes.cell(row=row, column=1).fill = warning_fill
            row += 1
        
        ws_faltantes.column_dimensions['A'].width = 50
        ws_faltantes.column_dimensions['B'].width = 25
    
    # Guardar
    wb.save(output_file)
    print(f"\n[OK] Informe de validación de métricas DIRECTAS guardado en: {output_file}")

def main():
    """Función principal."""
    print("=" * 80)
    print("VALIDACIÓN DE MÉTRICAS DIRECTAS (NO CALCULADAS)")
    print("Excluye: Subvenciones y métricas calculadas")
    print("=" * 80)
    
    try:
        # Cargar métricas extraídas
        print("\nBuscando archivo de métricas más reciente...")
        extracted_data = load_extracted_metrics()
        
        # Validar solo métricas directas
        print("Validando solo métricas DIRECTAS contra metodología oficial del INE...")
        validation_results = validate_direct_metrics(extracted_data)
        
        # Generar timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Guardar resultados JSON
        json_output = REPORTS_DIR / f"validacion_directas_{timestamp}.json"
        with open(json_output, 'w', encoding='utf-8') as f:
            json.dump(validation_results, f, ensure_ascii=False, indent=2)
        print(f"\n[OK] Resultados JSON guardados en: {json_output}")
        
        # Generar Excel
        excel_output = REPORTS_DIR / f"validacion_directas_{timestamp}.xlsx"
        generate_validation_report(validation_results, excel_output)
        
        # Mostrar resumen
        print("\n" + "=" * 80)
        print("RESUMEN DE VALIDACIÓN DE MÉTRICAS DIRECTAS")
        print("=" * 80)
        
        global_summary = validation_results['resumen_global']
        print(f"\nMétricas oficiales DIRECTAS INE: {global_summary['total_metricas_oficiales_directas']}")
        print(f"Métricas extraídas totales: {global_summary['total_metricas_extraidas']}")
        print(f"  - Validadas (directas): {global_summary['metricas_validadas']}")
        print(f"  - Calculadas detectadas: {global_summary['metricas_calculadas_detectadas']}")
        print(f"  - No identificadas: {global_summary['metricas_no_encontradas']}")
        print(f"\nPrecisión (sin calculadas): {global_summary['porcentaje_precision']:.1f}%")
        print(f"COBERTURA DE MÉTRICAS DIRECTAS: {global_summary['porcentaje_cobertura']:.1f}%")
        
        print("\nPor categoría:")
        for cat, data in validation_results['resumen_por_categoria'].items():
            status = "[OK]" if data['porcentaje_cobertura'] >= 80 else "[!]"
            print(f"  {status} {cat}: {data['encontradas_exactas']}/{data['total_oficiales']} ({data['porcentaje_cobertura']}%)")
        
        if validation_results['metricas_oficiales_faltantes']:
            print(f"\n[!] Métricas directas faltantes: {len(validation_results['metricas_oficiales_faltantes'])}")
            for item in validation_results['metricas_oficiales_faltantes'][:5]:
                print(f"    - {item['metrica']} ({item['categoria']})")
            if len(validation_results['metricas_oficiales_faltantes']) > 5:
                print(f"    ... y {len(validation_results['metricas_oficiales_faltantes']) - 5} más")
        
        print("\n[OK] Validación completada")
        print("[OK] Revise el archivo Excel para el informe detallado")
        
    except Exception as e:
        print(f"\n[ERROR] {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()